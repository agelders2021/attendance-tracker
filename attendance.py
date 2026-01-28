"""
SPDX-License-Identifier: GPL-3.0-or-later

Copyright (C) 2026 Al Gelders

This file is part of the attendance logging program

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

"""
ui.py - Main User Interface for the Training Attendance Tracker

This file contains the tkinter UI code for the training attendance tracking
application. Support methods are in ui_support.py and variables are in sv.py.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Optional
from datetime import datetime, timedelta
import os

import sv
import ui_support
import database
import config
from splash_screen import SplashScreen


class TrainingTrackerApp:
    """Main application class for the Training Attendance Tracker."""
    
    def __init__(self, root: tk.Tk):
        """Initialize the application.
        
        Args:
            root: The root tkinter window
        """
        self.root = root
        self.root.title("Training Attendance Tracker")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # Initialize variables
        self.vars = sv.init_variables(root)
        self.ui_state = ui_support.get_ui_state()
        self.db = database.get_db_manager()
        self.config = config.get_config_manager()
        
        # Store references to widgets we need to access later
        self.cert_tree: Optional[ttk.Treeview] = None
        self.members_tree: Optional[ttk.Treeview] = None
        self.attendance_tree: Optional[ttk.Treeview] = None
        self.last_name_combo: Optional[ttk.Combobox] = None
        self.location_combo: Optional[ttk.Combobox] = None
        self.primary_path_entry: Optional[ttk.Entry] = None
        self.secondary_path_entry: Optional[ttk.Entry] = None
        self.locations_listbox: Optional[tk.Listbox] = None
        self.location_entry: Optional[ttk.Entry] = None
        self.remove_location_btn: Optional[ttk.Button] = None
        self.certifications_listbox: Optional[tk.Listbox] = None
        self.certification_entry: Optional[ttk.Entry] = None
        self.remove_certification_btn: Optional[ttk.Button] = None
        self.date_entry = None  # TrainingDateEntry or fallback Entry
        self.description_frame: Optional[ttk.Frame] = None
        self.cert_date_entries: dict = {}
        self.cert_edit_popup: Optional[tk.Toplevel] = None
        
        # Flag to prevent duplicate name warning while loading member data
        self._loading_member_data = False
        
        # Flag to prevent session type change dialog while loading session data
        self._loading_session_data = False
        
        # Track which years have had attendance data modified during this session
        self._modified_years: set = set()
        
        # Create the main notebook (tab container)
        self._create_notebook()
        
    def _create_notebook(self):
        """Create the main notebook with tabs."""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create tab frames
        self.setup_tab = ttk.Frame(self.notebook)
        self.demographics_tab = ttk.Frame(self.notebook)
        self.training_sessions_tab = ttk.Frame(self.notebook)
        
        # Add tabs to notebook
        self.notebook.add(self.setup_tab, text="Setup")
        self.notebook.add(self.demographics_tab, text="Demographics")
        self.notebook.add(self.training_sessions_tab, text="Sessions")
        
        # Build each tab
        self._build_setup_tab()
        self._build_demographics_tab()
        self._build_training_sessions_tab()
        
        # Track current tab for change detection
        self.current_tab_index = 0
        
        # Bind tab change event
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        
        # Load configuration and initialize database connection
        self._load_config_and_initialize()
        
        # Bind window close to save config
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        
    def _on_tab_changed(self, event):
        """Handle tab change event.
        
        Args:
            event: The tkinter event
        """
        new_tab_index = self.notebook.index(self.notebook.select())
        self.current_tab_index = new_tab_index
        
    def _get_current_session_data(self) -> dict:
        """Get current session form data for change tracking.
        
        Returns:
            Dictionary of current session field values
        """
        return {
            'location': self.vars.session_location.get(),
            'date': self.vars.session_date.get(),
            'type': self.vars.session_type.get(),
            'description': self.vars.session_description.get(),
        }
        
    def _load_config_and_initialize(self):
        """Load configuration from file and initialize database if path exists."""
        # Load saved paths
        primary_path = self.config.primary_storage_folder
        secondary_path = self.config.secondary_backup_folder
        excel_path = self.config.excel_participation_folder
        
        database_exists = False
        
        if primary_path:
            self.vars.primary_storage_folder.set(primary_path)
            
            # Set database path and try to connect
            db_path = os.path.join(primary_path, "training_tracker.db")
            self.db.set_db_path(db_path)
            
            if self.db.database_exists():
                database_exists = True
                self._load_data_from_database()
                
                # Check if backup is newer than database
                if secondary_path:
                    self._check_backup_vs_database(secondary_path)
                
        if secondary_path:
            self.vars.secondary_backup_folder.set(secondary_path)
            
        if excel_path:
            self.vars.excel_participation_folder.set(excel_path)
            
        # Load email settings
        self.vars.smtp_server.set(self.config.smtp_server or "smtp.gmail.com")
        self.vars.smtp_port.set(str(self.config.smtp_port or 587))
        self.vars.smtp_username.set(self.config.smtp_username or "")
        self.vars.sender_email.set(self.config.sender_email or "")
        self.vars.sender_password.set(self.config.sender_password or "")
        self.vars.smtp_encryption.set(self.config.smtp_encryption or "TLS")
        # Always default to 1 month - don't remember from previous session
        self.vars.pdf_months_to_review.set(1)
            
        # Restore window geometry if saved
        geometry = self.config.window_geometry
        if geometry:
            try:
                self.root.geometry(geometry)
            except tk.TclError:
                pass
                
        # If database exists, go directly to Sessions tab (index 2)
        if database_exists:
            self.notebook.select(2)
                
    def _check_backup_vs_database(self, backup_folder: str):
        """Check if most recent backup is newer than database and offer to restore.
        
        Args:
            backup_folder: Path to backup folder
        """
        most_recent = config.get_most_recent_backup(backup_folder)
        if not most_recent:
            return
            
        backup_path, backup_timestamp = most_recent
        
        # Get database last modified time
        db_time = self.db.get_last_modified_time()
        
        if db_time is None:
            return
            
        # Compare timestamps (backup_timestamp is already a datetime)
        if backup_timestamp > db_time:
            response = messagebox.askyesno(
                "Newer Backup Found",
                f"A backup from {backup_timestamp.strftime('%Y-%m-%d %H:%M:%S')} "
                f"is newer than the database.\n\n"
                f"Database last modified: {db_time.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                "Would you like to restore from the backup?"
            )
            
            if response:
                self._restore_from_backup_file(backup_path)
                
    def _on_closing(self):
        """Handle window closing - create backup and save configuration."""
        # Generate Excel participation file if folder is set and data was modified
        excel_folder = self.vars.excel_participation_folder.get()
        if self.db.database_exists() and excel_folder and self._modified_years:
            self._generate_participation_excel(excel_folder)
        
        # Create backup only if data was modified (indicated by _modified_years not empty)
        backup_folder = self.vars.secondary_backup_folder.get()
        if self.db.database_exists() and backup_folder and self._modified_years:
            self._create_exit_backup(backup_folder)
            
            # Check if we should clean up old backups (once per month)
            self._cleanup_old_backups(backup_folder)
        
        # Save pdf_months_to_review BEFORE sending emails so it uses the current spinbox value
        self.config.pdf_months_to_review = self.vars.pdf_months_to_review.get()
            
        # Send monthly emails if it's a new month
        if self.db.database_exists():
            self._send_monthly_emails_if_needed()
        
        # Save current settings to config (pdf_months_to_review already saved above)
        self.config.primary_storage_folder = self.vars.primary_storage_folder.get()
        self.config.secondary_backup_folder = self.vars.secondary_backup_folder.get()
        self.config.excel_participation_folder = self.vars.excel_participation_folder.get()
        self.config.smtp_server = self.vars.smtp_server.get()
        try:
            self.config.smtp_port = int(self.vars.smtp_port.get())
        except ValueError:
            self.config.smtp_port = 587
        self.config.smtp_username = self.vars.smtp_username.get()
        self.config.sender_email = self.vars.sender_email.get()
        self.config.sender_password = self.vars.sender_password.get()
        self.config.smtp_encryption = self.vars.smtp_encryption.get()
        self.config.window_geometry = self.root.geometry()
        self.config.save()
        
        # Close database connection
        self.db.close()
        
        # Destroy window
        self.root.destroy()
        
    def _create_exit_backup(self, backup_folder: str):
        """Create a backup on exit.
        
        Args:
            backup_folder: Path to backup folder
        """
        try:
            data = self.db.export_all_data()
            backup_path = config.create_backup(backup_folder, data)
            if backup_path:
                print(f"Backup created: {backup_path}")
        except Exception as e:
            print(f"Error creating backup: {e}")
            
    def _cleanup_old_backups(self, backup_folder: str):
        """Clean up old backups if it's been more than a month since last cleanup.
        
        Args:
            backup_folder: Path to backup folder
        """
        last_cleanup = self.config.last_backup_cleanup_date
        
        if not config.should_cleanup_backups(last_cleanup):
            return
            
        old_backups = config.get_old_backups(backup_folder, months=6)
        
        if not old_backups:
            return
            
        # Ask user before deleting
        response = messagebox.askyesno(
            "Clean Up Old Backups",
            f"Found {len(old_backups)} backup(s) older than 6 months.\n\n"
            "Would you like to delete them to free up space?"
        )
        
        if response:
            deleted = 0
            for filepath, _ in old_backups:
                if config.delete_backup(filepath):
                    deleted += 1
                
        # Update last cleanup date
        self.config.last_backup_cleanup_date = datetime.now().isoformat()
        self.config.save()
        
    def _generate_participation_excel(self, excel_folder: str):
        """Generate Participation-{year}.xlsx files for modified years only.
        
        Args:
            excel_folder: Path to the Excel output folder
        """
        # Only generate for years that were modified during this session
        if not self._modified_years:
            return
            
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from collections import defaultdict
            import calendar
            
            # Get all sessions and members
            sessions = self.db.get_all_sessions()
            members = self.db.get_all_members()
            
            # Defensive checks - ensure we have valid lists
            if sessions is None or members is None:
                return
            if not sessions or not members:
                return
            
            # Group sessions by year, then by month
            sessions_by_year = defaultdict(lambda: defaultdict(list))
            for session in sessions:
                date_str = session.get('date', '')
                try:
                    session_date = datetime.strptime(date_str, "%m/%d/%Y")
                    year = session_date.year
                    month = session_date.month
                    # Store session with parsed date for sorting
                    session['_parsed_date'] = session_date
                    sessions_by_year[year][month].append(session)
                except ValueError:
                    continue
            
            # Define fill colors for session types (header cells only)
            fills = {
                'Weeknight Training': PatternFill('solid', fgColor='ADD8E6'),   # Light blue
                'Weekend Training': PatternFill('solid', fgColor='FFFF00'),   # Yellow
                'Mission': PatternFill('solid', fgColor='FF6B6B'),   # Light red
                'Team Meeting': PatternFill('solid', fgColor='DDA0DD'),   # Plum
                'Other': PatternFill('solid', fgColor='90EE90'),     # Light green
            }
            
            # Warning row styling (red background, white bold text)
            warning_fill = PatternFill('solid', fgColor='FF0000')
            warning_font = Font(bold=True, color='FFFFFF')
            
            # Generate spreadsheet for each modified year
            for year in self._modified_years:
                if year not in sessions_by_year:
                    continue
                    
                filename = f"Participation-{year}.xlsx"
                filepath = os.path.join(excel_folder, filename)
                backup_path = os.path.join(excel_folder, f"Participation-{year}-back.xlsx")
                
                # Move current file to backup if it exists
                if os.path.exists(filepath):
                    if os.path.exists(backup_path):
                        os.remove(backup_path)
                    os.rename(filepath, backup_path)
                
                # Create workbook
                wb = Workbook()
                # Remove default sheet
                wb.remove(wb.active)
                
                # Sort months in chronological order (January first)
                year_data = sessions_by_year.get(year)
                if not year_data:
                    continue
                sorted_months = sorted(year_data.keys())
                if not sorted_months:
                    continue
                
                for month in sorted_months:
                    month_sessions = sessions_by_year[year][month]
                    
                    # Skip if no sessions for this month
                    if not month_sessions:
                        continue
                    
                    # Sort sessions within month by date (chronological - earlier dates on left)
                    month_sessions.sort(key=lambda s: s['_parsed_date'])
                    
                    month_name = calendar.month_name[month]
                    sheet_name = f"{month_name}-{year}"
                    
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Row 1: Warning header - spans across all columns
                    warning_text = "***** Auto Generated -- Do not edit! *****"
                    total_cols = len(month_sessions) + 1  # Member column + session columns
                    
                    warning_cell = ws.cell(row=1, column=1, value=warning_text)
                    warning_cell.font = warning_font
                    warning_cell.fill = warning_fill
                    warning_cell.alignment = Alignment(horizontal='center')
                    
                    # Merge cells for warning row if there are multiple columns
                    if total_cols > 1:
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                    
                    # Row 2: Header row - first column is "Member"
                    ws.cell(row=2, column=1, value="Member")
                    ws.cell(row=2, column=1).font = Font(bold=True)
                    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
                    
                    # Add session headers: {M/D} {location} - color coded by type
                    # Drop year since it's in the sheet title
                    for col, session in enumerate(month_sessions, start=2):
                        parsed_date = session['_parsed_date']
                        # Format as M/D (no year, no leading zeros)
                        date_short = f"{parsed_date.month}/{parsed_date.day}"
                        location = session.get('location', '')
                        session_type = session.get('type', 'Weekend Training')
                        header_text = f"{date_short} {location}"
                        
                        cell = ws.cell(row=2, column=col, value=header_text)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Apply color to HEADER ONLY based on session type
                        if session_type in fills:
                            cell.fill = fills[session_type]
                    
                    # Add member rows (NO coloring on data cells) - starting at row 3
                    for row, member in enumerate(members, start=3):
                        # Member name
                        name = f"{member.get('last_name', '')}, {member.get('first_name', '')}"
                        ws.cell(row=row, column=1, value=name)
                        
                        # Attendance for each session
                        member_id = member.get('id')
                        for col, session in enumerate(month_sessions, start=2):
                            session_id = session.get('id')
                            
                            # Get attendance status
                            attended = self.db.get_attendance_status(session_id, member_id)
                            cell = ws.cell(row=row, column=col, value="Yes" if attended else "No")
                            cell.alignment = Alignment(horizontal='center')
                    
                    # Auto-adjust column widths
                    ws.column_dimensions['A'].width = 25
                    for col in range(2, len(month_sessions) + 2):
                        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 18
                    
                    # Protect the worksheet to make it read-only in Excel
                    # Users can view but cannot edit without unprotecting
                    # No password set - just a visual deterrent
                    ws.protection.sheet = True
                    ws.protection.enable()
                
                # Save workbook
                wb.save(filepath)
            
        except ImportError:
            pass
        except Exception as e:
            print(f"Error generating Excel: {e}")
            
    def _send_monthly_emails_if_needed(self):
        """Send monthly attendance emails if it's a new month."""
        current_month = datetime.now().strftime("%Y-%m")
        last_email_month = self.config.last_email_month
        
        # ===DEBUG=== Send emails every exit for testing
        # print("DEBUG: _send_monthly_emails_if_needed called")
        # print(f"DEBUG: Current month: {current_month}, Last email month: {last_email_month}")
        # self._send_attendance_emails_to_all()
        # return
        # ===END DEBUG===
        
        if current_month != last_email_month:
            self._send_attendance_emails_to_all()
            self.config.last_email_month = current_month
            self.config.save()
            
    def _show_attendance_review_dialog(self, members: list):
        """Show a dialog to select members and date range for PDF generation.
        
        Args:
            members: List of member dictionaries
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("Review Attendance")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center on parent
        dialog.geometry(f"+{self.root.winfo_x() + 200}+{self.root.winfo_y() + 50}")
        
        # Date range section
        date_frame = ttk.LabelFrame(dialog, text="Date Range", padding=5)
        date_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # Start Date
        start_row = ttk.Frame(date_frame)
        start_row.pack(fill=tk.X, pady=2)
        
        ttk.Label(start_row, text="Start Date:", width=10).pack(side=tk.LEFT)
        
        # Use DateEntry if available, otherwise regular entry
        if ui_support.TKCALENDAR_AVAILABLE:
            from tkcalendar import DateEntry
            start_date_var = tk.StringVar()
            start_date_entry = DateEntry(start_row, textvariable=start_date_var, 
                                          date_pattern='mm/dd/yyyy', width=12)
            # Default to 1 month ago
            default_start = datetime.now() - timedelta(days=30)
            start_date_entry.set_date(default_start)
        else:
            start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime("%m/%d/%Y"))
            start_date_entry = ttk.Entry(start_row, textvariable=start_date_var, width=15)
        start_date_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # End Date
        end_row = ttk.Frame(date_frame)
        end_row.pack(fill=tk.X, pady=2)
        
        ttk.Label(end_row, text="End Date:", width=10).pack(side=tk.LEFT)
        
        if ui_support.TKCALENDAR_AVAILABLE:
            end_date_var = tk.StringVar()
            end_date_entry = DateEntry(end_row, textvariable=end_date_var, 
                                        date_pattern='mm/dd/yyyy', width=12)
            # Default to today
            end_date_entry.set_date(datetime.now())
        else:
            end_date_var = tk.StringVar(value=datetime.now().strftime("%m/%d/%Y"))
            end_date_entry = ttk.Entry(end_row, textvariable=end_date_var, width=15)
        end_date_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # Member selection section
        member_frame = ttk.LabelFrame(dialog, text="Select Members", padding=5)
        member_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Listbox with scrollbar for multi-selection
        list_frame = ttk.Frame(member_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, selectmode=tk.EXTENDED, height=12)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Sort members by status first, then by name
        sorted_members = ui_support.sort_members_by_status_then_name(members, sort_by_first=False)
        
        # Populate listbox with members, track indices by status
        member_map = {}
        member_indices = []  # List of (index, member) tuples
        status_indices = {'Member': [], 'Candidate': [], 'Guest': []}
        
        for i, member in enumerate(sorted_members):
            status = member.get('member_status', 'Member')
            display_name = f"{member.get('last_name', '')}, {member.get('first_name', '')} ({status})"
            listbox.insert(tk.END, display_name)
            member_map[display_name] = member
            member_indices.append((i, member))
            if status in status_indices:
                status_indices[status].append(i)
        
        # Selection button states (for toggle functionality)
        button_states = {
            'all': False,  # False = Select All, True = Deselect All
            'Member': False,
            'Candidate': False,
            'Guest': False
        }
        
        # Selection buttons frame
        select_frame = ttk.Frame(member_frame)
        select_frame.pack(fill=tk.X, pady=5)
        
        def toggle_all():
            if not button_states['all']:
                listbox.select_set(0, tk.END)
                select_all_btn.config(text="Deselect All")
                button_states['all'] = True
            else:
                listbox.select_clear(0, tk.END)
                select_all_btn.config(text="Select All")
                button_states['all'] = False
                # Reset all status buttons
                for status in ['Member', 'Candidate', 'Guest']:
                    button_states[status] = False
                select_members_btn.config(text="Select Members")
                select_candidates_btn.config(text="Select Candidates")
                select_guests_btn.config(text="Select Guests")
        
        def toggle_status(status, button):
            indices = status_indices.get(status, [])
            if not button_states[status]:
                for idx in indices:
                    listbox.select_set(idx)
                button.config(text=f"Deselect {status}s")
                button_states[status] = True
            else:
                for idx in indices:
                    listbox.select_clear(idx)
                button.config(text=f"Select {status}s")
                button_states[status] = False
        
        select_all_btn = ttk.Button(select_frame, text="Select All", command=toggle_all)
        select_all_btn.pack(side=tk.LEFT, padx=5)
        
        select_members_btn = ttk.Button(select_frame, text="Select Members", 
                                         command=lambda: toggle_status('Member', select_members_btn))
        select_members_btn.pack(side=tk.LEFT, padx=5)
        
        select_candidates_btn = ttk.Button(select_frame, text="Select Candidates", 
                                            command=lambda: toggle_status('Candidate', select_candidates_btn))
        select_candidates_btn.pack(side=tk.LEFT, padx=5)
        
        select_guests_btn = ttk.Button(select_frame, text="Select Guests", 
                                        command=lambda: toggle_status('Guest', select_guests_btn))
        select_guests_btn.pack(side=tk.LEFT, padx=5)
        
        # Button frame
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def on_execute():
            # Validate date range
            start_str = start_date_var.get().strip()
            end_str = end_date_var.get().strip()
            
            if not start_str or not end_str:
                messagebox.showwarning("Date Required", "Please enter both Start Date and End Date.")
                return
            
            # Validate dates
            is_valid, result = ui_support.validate_date(start_str)
            if not is_valid:
                messagebox.showwarning("Invalid Start Date", result)
                return
            start_str = result
            
            is_valid, result = ui_support.validate_date(end_str)
            if not is_valid:
                messagebox.showwarning("Invalid End Date", result)
                return
            end_str = result
            
            # Parse dates
            try:
                start_date = datetime.strptime(start_str, "%m/%d/%Y")
                end_date = datetime.strptime(end_str, "%m/%d/%Y")
            except ValueError as e:
                messagebox.showwarning("Date Error", f"Error parsing dates: {e}")
                return
            
            if start_date > end_date:
                messagebox.showwarning("Date Range Error", "Start Date must be before or equal to End Date.")
                return
            
            # Validate member selection
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select at least one member.")
                return
            
            # Get selected members
            selected_members = []
            for idx in selection:
                display_name = listbox.get(idx)
                member = member_map.get(display_name)
                if member:
                    selected_members.append(member)
            
            dialog.destroy()
            
            # Generate PDF for selected members with date range
            self._generate_pdf_for_date_range(selected_members, start_date, end_date)
        
        def on_cancel():
            dialog.destroy()
        
        ttk.Button(btn_frame, text="Execute", command=on_execute).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.LEFT, padx=5)
        
    def _generate_pdf_for_date_range(self, selected_members: list, start_date: datetime, end_date: datetime):
        """Generate PDF for the selected members within the specified date range.
        
        Args:
            selected_members: List of member dictionaries to include
            start_date: Start date for session filtering
            end_date: End date for session filtering
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            import tempfile
            import subprocess
            import sys
            
            if not selected_members:
                return
            
            all_sessions = self.db.get_all_sessions()
            
            # Filter sessions to the specified date range
            sessions_in_range = []
            for session in all_sessions:
                try:
                    session_date = datetime.strptime(session.get('date', ''), "%m/%d/%Y")
                    if start_date <= session_date <= end_date:
                        sessions_in_range.append(session)
                except ValueError:
                    continue
            
            if not sessions_in_range:
                messagebox.showinfo("No Sessions", 
                                    f"No sessions found between {start_date.strftime('%m/%d/%Y')} and {end_date.strftime('%m/%d/%Y')}.")
                return
            
            # Count totals for the date range
            total_qualifying = sum(1 for s in sessions_in_range if s.get('type') == 'Weekend Training')
            total_optional = sum(1 for s in sessions_in_range if s.get('type') == 'Weeknight Training')
            total_mission = sum(1 for s in sessions_in_range if s.get('type') == 'Mission')
            total_team_meeting = sum(1 for s in sessions_in_range if s.get('type') == 'Team Meeting')
            total_other = sum(1 for s in sessions_in_range if s.get('type') == 'Other')
            
            # Create temp file for combined PDF
            fd, pdf_path = tempfile.mkstemp(suffix='.pdf')
            os.close(fd)
            
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Format date range for display
            date_range_str = f"{start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"
            
            # Sort members by last name, first name
            sorted_members = sorted(selected_members, key=lambda m: (m.get('last_name', '').lower(), m.get('first_name', '').lower()))
            
            # ============================================================
            # Summary Page - All Members Overview
            # ============================================================
            story.append(Paragraph("505 SAR Dogs Attendance Summary", styles['Title']))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Period: {date_range_str}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Session totals sentences
            story.append(Paragraph(f"Total Weekend Training sessions: {total_qualifying}", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Total Weeknight Training sessions: {total_optional}", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Total Missions: {total_mission}", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Total Team Meetings: {total_team_meeting}", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Total Other sessions: {total_other}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Helper function to format count with percentage
            def format_with_percent(attended, total):
                if total == 0:
                    return "0 (0%)"
                percent = int(round(attended / total * 100))
                return f"{attended} ({percent}%)"
            
            # Build summary table data - calculate attendance for each member
            summary_table_data = [['Member', 'Weekend\nSessions', 'Weeknight\nSessions', 'Missions', 'Meetings', 'Other']]
            
            for member in sorted_members:
                member_id = member.get('id')
                member_name = f"{member.get('last_name', '')}, {member.get('first_name', '')}"
                
                # Calculate attendance for each session type
                qualifying_attended = 0
                optional_attended = 0
                mission_attended = 0
                team_meeting_attended = 0
                other_attended = 0
                
                for session in sessions_in_range:
                    session_id = session.get('id')
                    attended = self.db.get_attendance_status(session_id, member_id)
                    session_type = session.get('type', '')
                    if attended:
                        if session_type == 'Weekend Training':
                            qualifying_attended += 1
                        elif session_type == 'Weeknight Training':
                            optional_attended += 1
                        elif session_type == 'Mission':
                            mission_attended += 1
                        elif session_type == 'Team Meeting':
                            team_meeting_attended += 1
                        elif session_type == 'Other':
                            other_attended += 1
                
                summary_table_data.append([
                    member_name,
                    format_with_percent(qualifying_attended, total_qualifying),
                    format_with_percent(optional_attended, total_optional),
                    format_with_percent(mission_attended, total_mission),
                    format_with_percent(team_meeting_attended, total_team_meeting),
                    format_with_percent(other_attended, total_other)
                ])
            
            # Create summary table with repeatRows=1 to repeat header on page breaks
            summary_table = Table(summary_table_data, colWidths=[150, 70, 70, 60, 60, 60], repeatRows=1)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Member names left-aligned
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Numbers centered
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(summary_table)
            
            # ============================================================
            # Individual Member Pages
            # ============================================================
            for member in sorted_members:
                story.append(PageBreak())
                    
                member_id = member.get('id')
                member_name = f"{member.get('first_name', '')} {member.get('last_name', '')}"
                
                # Calculate attendance for each session type within the date range
                qualifying_attended = 0
                optional_attended = 0
                mission_attended = 0
                team_meeting_attended = 0
                other_attended = 0
                
                for session in sessions_in_range:
                    session_id = session.get('id')
                    attended = self.db.get_attendance_status(session_id, member_id)
                    session_type = session.get('type', '')
                    if attended:
                        if session_type == 'Weekend Training':
                            qualifying_attended += 1
                        elif session_type == 'Weeknight Training':
                            optional_attended += 1
                        elif session_type == 'Mission':
                            mission_attended += 1
                        elif session_type == 'Team Meeting':
                            team_meeting_attended += 1
                        elif session_type == 'Other':
                            other_attended += 1
                
                # Title
                story.append(Paragraph(f"505 SAR Dogs Attendance Record for {member_name}", styles['Title']))
                story.append(Spacer(1, 12))
                
                # Date range subtitle
                story.append(Paragraph(f"Period: {date_range_str}", styles['Normal']))
                story.append(Spacer(1, 12))
                
                # Summary lines
                story.append(Paragraph(f"Attended {qualifying_attended} out of {total_qualifying} weekend training sessions.", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"Attended {optional_attended} out of {total_optional} weeknight training sessions.", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"Attended {mission_attended} out of {total_mission} missions.", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"Attended {team_meeting_attended} out of {total_team_meeting} team meetings.", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"Attended {other_attended} out of {total_other} other sessions.", styles['Normal']))
                story.append(Spacer(1, 20))
                
                # Build table with sessions in range
                if sessions_in_range:
                    sessions_sorted = sorted(sessions_in_range, key=lambda s: datetime.strptime(s.get('date', '01/01/2000'), "%m/%d/%Y"))
                    
                    table_data = [[f'Sessions from {date_range_str}', '', '', '']]
                    table_data.append(['Date', 'Location', 'Type', 'Attended'])
                    
                    for session in sessions_sorted:
                        session_id = session.get('id')
                        attended = self.db.get_attendance_status(session_id, member_id)
                        attended_str = "Yes" if attended else "No"
                        
                        # Convert old session type names to new names for display
                        # For 'Other' type, use the description instead
                        session_type = session.get('type', '')
                        if session_type == 'Qualifying Training':
                            session_type = 'Weekend Training'
                        elif session_type == 'Optional Training':
                            session_type = 'Weeknight Training'
                        elif session_type == 'Other':
                            description = session.get('description', '').strip()
                            if description:
                                session_type = description
                        
                        table_data.append([
                            session.get('date', ''),
                            session.get('location', ''),
                            session_type,
                            attended_str
                        ])
                    
                    table = Table(table_data, colWidths=[80, 150, 100, 60], repeatRows=2)
                    table.setStyle(TableStyle([
                        ('SPAN', (0, 0), (-1, 0)),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('TOPPADDING', (0, 0), (-1, 0), 8),
                        ('BACKGROUND', (0, 1), (-1, 1), colors.grey),
                        ('TEXTCOLOR', (0, 1), (-1, 1), colors.whitesmoke),
                        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 1), (-1, 1), 10),
                        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
                        ('BACKGROUND', (0, 2), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    
                    story.append(table)
            
            # Build PDF
            doc.build(story)
            
            # Open with system default PDF viewer
            if sys.platform == 'win32':
                os.startfile(pdf_path)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', pdf_path])
            else:  # Linux
                subprocess.run(['xdg-open', pdf_path])
                
        except ImportError as e:
            print(f"Required library not installed for PDF: {e}")
            messagebox.showerror("Missing Library", f"Required library not installed: {e}")
        except Exception as e:
            print(f"Error generating PDF preview: {e}")
            messagebox.showerror("PDF Error", f"Error generating PDF: {e}")
            
    def _send_attendance_emails_to_all(self):
        """Send attendance PDF emails to all members with email addresses."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            import smtplib
            import ssl
            from email.mime.multipart import MIMEMultipart
            from email.mime.base import MIMEBase
            from email.mime.text import MIMEText
            from email import encoders
            import tempfile
            import subprocess
            
            # Check for email configuration
            sender_email = self.config.sender_email
            sender_password = self.config.sender_password
            smtp_server = self.config.smtp_server
            smtp_port = self.config.smtp_port
            smtp_username = self.config.smtp_username or sender_email  # Default to sender_email if no username
            smtp_encryption = self.config.smtp_encryption
            
            if not sender_email or not sender_password:
                print("Email not configured - skipping email send. Set sender_email and sender_password in config.")
                return
            
            members = self.db.get_all_members()
            
            # Look up the sender's member name based on their email
            sender_member_name = "Secretary"  # Default fallback
            for m in members:
                if m.get('email', '').strip().lower() == sender_email.strip().lower():
                    sender_member_name = f"{m.get('first_name', '')} {m.get('last_name', '')}".strip()
                    break
                elif m.get('alternate_email', '').strip().lower() == sender_email.strip().lower():
                    sender_member_name = f"{m.get('first_name', '')} {m.get('last_name', '')}".strip()
                    break
            
            # Get the configured number of months to review for table
            months_to_review = self.config.pdf_months_to_review
            
            # Get sessions from last 6 months (for weekend attendance count)
            cutoff_6mo = datetime.now() - timedelta(days=180)
            # Get sessions for the table based on configured months
            cutoff_table = datetime.now() - timedelta(days=months_to_review * 30) if months_to_review > 0 else None
            all_sessions = self.db.get_all_sessions()
            
            # Filter sessions to last 6 months and table period
            sessions_6mo = []
            sessions_table = []
            for session in all_sessions:
                try:
                    # NOTE: get_all_sessions renames session_date->date
                    session_date = datetime.strptime(session.get('date', ''), "%m/%d/%Y")
                    if session_date >= cutoff_6mo:
                        sessions_6mo.append(session)
                    if cutoff_table and session_date >= cutoff_table:
                        sessions_table.append(session)
                except ValueError:
                    continue
            
            # DEBUG: Print session filtering results
            # print(f"DEBUG: Email - Total sessions: {len(all_sessions)}, Last 6mo: {len(sessions_6mo)}, Last 3mo: {len(sessions_3mo)}")
            # if sessions_6mo:
            #     print(f"DEBUG: Email - First 6mo session: {sessions_6mo[0]}")
            
            if not sessions_6mo:
                # DEBUG: print("DEBUG: Email - No sessions in last 6 months for email")
                return
                
            # Count total qualifying training sessions in last 6 months
            # NOTE: get_all_sessions renames session_type->type
            total_qualifying_sessions = sum(1 for s in sessions_6mo if s.get('type') == 'Weekend Training')
            total_optional_sessions = sum(1 for s in sessions_6mo if s.get('type') == 'Weeknight Training')
            total_mission_sessions = sum(1 for s in sessions_6mo if s.get('type') == 'Mission')
            total_team_meeting_sessions = sum(1 for s in sessions_6mo if s.get('type') == 'Team Meeting')
            total_other_sessions = sum(1 for s in sessions_6mo if s.get('type') == 'Other')
            # DEBUG: print(f"DEBUG: Email - Total qualifying training sessions in 6mo period: {total_qualifying_sessions}")
            
            # ===DEBUG=== Generate PDF preview and send emails
            # print("DEBUG: Email - Generating PDFs and sending emails")
            
            # Connect to SMTP server once for all emails
            # DEBUG: print(f"DEBUG: Email - Connecting to SMTP server: {smtp_server}:{smtp_port}")
            # DEBUG: print(f"DEBUG: Email - Encryption type: {smtp_encryption}")
            # DEBUG: print(f"DEBUG: Email - Username: {smtp_username}")
            # DEBUG: print(f"DEBUG: Email - Sender: {sender_email}")
            
            server = None
            try:
                if smtp_encryption == "SSL":
                    # DEBUG: print("DEBUG: Email - Using SSL connection")
                    context = ssl.create_default_context()
                    server = smtplib.SMTP_SSL(smtp_server, smtp_port, context=context)
                else:
                    # DEBUG: print("DEBUG: Email - Using plain SMTP connection")
                    server = smtplib.SMTP(smtp_server, smtp_port)
                    # DEBUG: server.set_debuglevel(1)  # Enable SMTP debug output
                    
                    if smtp_encryption == "TLS":
                        # DEBUG: print("DEBUG: Email - Starting TLS")
                        server.starttls()
                        # DEBUG: print(f"DEBUG: Email - STARTTLS result: {tls_result}")
                    elif smtp_encryption == "STARTTLS":
                        # DEBUG: print("DEBUG: Email - Starting STARTTLS")
                        server.starttls()
                        # DEBUG: print(f"DEBUG: Email - STARTTLS result: {tls_result}")
                        
                # DEBUG: print("DEBUG: Email - Attempting login...")
                server.login(smtp_username, sender_password)
                # DEBUG: print(f"DEBUG: Email - Login result: {login_result}")
                
            except smtplib.SMTPAuthenticationError as e:
                # DEBUG: print(f"DEBUG: Email - Authentication failed: code={e.smtp_code}, msg={e.smtp_error}")
                if server:
                    server.quit()
                return
            except smtplib.SMTPException as e:
                # DEBUG: print(f"DEBUG: Email - SMTP error: {type(e).__name__}: {e}")
                # if hasattr(e, 'smtp_code'):
                #     print(f"DEBUG: Email - SMTP code: {e.smtp_code}")
                # if hasattr(e, 'smtp_error'):
                #     print(f"DEBUG: Email - SMTP error msg: {e.smtp_error}")
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                return
            except Exception as e:
                # DEBUG: print(f"DEBUG: Email - Connection error: {type(e).__name__}: {e}")
                if server:
                    try:
                        server.quit()
                    except:
                        pass
                return
            
            emails_sent = 0
            for member in members:
                email = member.get('email', '').strip()
                if not email:
                    continue
                    
                member_id = member.get('id')
                member_name = f"{member.get('first_name', '')} {member.get('last_name', '')}"
                # DEBUG: print(f"DEBUG: Email - Processing member: {member_name} <{email}>")
                
                # Generate PDF for this member
                # Pass table sessions for table, 6mo sessions for counts, and months_to_review
                pdf_path = self._generate_member_attendance_pdf(
                    member, sessions_table, sessions_6mo, 
                    total_qualifying_sessions, total_optional_sessions,
                    total_mission_sessions, total_team_meeting_sessions, total_other_sessions,
                    months_to_review
                )
                
                if pdf_path:
                    try:
                        # Create email message
                        msg = MIMEMultipart()
                        msg['From'] = sender_email
                        msg['To'] = email
                        msg['Subject'] = "505 SAR Dogs Attendance Record"
                        
                        # Email body
                        month_text = "month" if months_to_review == 1 else "months"
                        if months_to_review > 0:
                            table_info = f"for the last {months_to_review} {month_text}"
                        else:
                            table_info = "(summary only)"
                        body = (
                            f"To: {member_name},\n"
                            f"Re: Training\n\n"
                            f"Please find attached your training attendance record {table_info}.\n\n"
                            f"Contact me (Secretary) by replying to this email if you think there is an error.\n\n"
                            f"Thank you,\n"
                            f"{sender_member_name}"
                        )
                        msg.attach(MIMEText(body, 'plain'))
                        
                        # Attach PDF
                        with open(pdf_path, 'rb') as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename="Attendance_Record_{member.get("last_name", "")}.pdf"'
                        )
                        msg.attach(part)
                        
                        # Send email
                        # DEBUG: print(f"DEBUG: Email - Sending email to {email}...")
                        server.sendmail(sender_email, email, msg.as_string())
                        # DEBUG: print(f"DEBUG: Email - Send result: {send_result}")
                        emails_sent += 1
                        # DEBUG: print(f"DEBUG: Email - Successfully sent to {email}")
                        
                    except smtplib.SMTPRecipientsRefused as e:
                        # DEBUG: print(f"DEBUG: Email - Recipients refused for {email}: {e.recipients}")
                        pass
                    except smtplib.SMTPException as e:
                        # DEBUG: print(f"DEBUG: Email - SMTP error sending to {email}: {type(e).__name__}: {e}")
                        # if hasattr(e, 'smtp_code'):
                        #     print(f"DEBUG: Email - SMTP code: {e.smtp_code}")
                        pass
                    except Exception as e:
                        # DEBUG: print(f"DEBUG: Email - Error sending to {email}: {type(e).__name__}: {e}")
                        pass
                    
                    # Clean up temp file
                    try:
                        os.remove(pdf_path)
                        # DEBUG: print(f"DEBUG: Email - Cleaned up temp file: {pdf_path}")
                    except Exception as e:
                        # DEBUG: print(f"DEBUG: Email - Could not remove temp file: {e}")
                        pass
                # DEBUG: else:
                #     print(f"DEBUG: Email - PDF generation failed for {member_name}")
            
            # Close SMTP connection
            try:
                server.quit()
                # DEBUG: print("DEBUG: Email - SMTP connection closed")
            except Exception as e:
                # DEBUG: print(f"DEBUG: Email - Error closing SMTP: {e}")
                pass
                
            # DEBUG: print(f"DEBUG: Email - Total emails sent: {emails_sent}")
            # ===END DEBUG===
                        
        except ImportError as e:
            print(f"Required library not installed for email: {e}")
        except Exception as e:
            print(f"Error sending emails: {e}")
            
    def _generate_member_attendance_pdf(self, member: dict, sessions_table: list, sessions_6mo: list, 
                                          total_qualifying: int, total_optional: int,
                                          total_mission: int, total_team_meeting: int, total_other: int,
                                          months_to_review: int = 1) -> str:
        """Generate attendance PDF for a single member.
        
        Args:
            member: Member dictionary
            sessions_table: List of session dictionaries for table (based on months_to_review)
            sessions_6mo: List of session dictionaries for last 6 months (for counts)
            total_qualifying: Total number of qualifying training sessions in last 6 months
            total_optional: Total number of optional training sessions in last 6 months
            total_mission: Total number of missions in last 6 months
            total_team_meeting: Total number of team meetings in last 6 months
            total_other: Total number of other sessions in last 6 months
            months_to_review: Number of months to show in the table (0 = no table)
            
        Returns:
            Path to generated PDF file, or None on error
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            import tempfile
            
            member_id = member.get('id')
            member_name = f"{member.get('first_name', '')} {member.get('last_name', '')}"
            
            # DEBUG: Print member info
            # print(f"DEBUG: PDF - Generating PDF for {member_name} (ID: {member_id})")
            # print(f"DEBUG: PDF - Sessions for table: {len(sessions_table)}, Sessions for count (6mo): {len(sessions_6mo)}, Total qualifying: {total_qualifying}")
            
            # Calculate attendance for each session type from 6 month data
            qualifying_attended = 0
            optional_attended = 0
            mission_attended = 0
            team_meeting_attended = 0
            other_attended = 0
            
            for session in sessions_6mo:
                session_id = session.get('id')
                attended = self.db.get_attendance_status(session_id, member_id)
                session_type = session.get('type', '')
                if attended:
                    if session_type == 'Weekend Training':
                        qualifying_attended += 1
                    elif session_type == 'Weeknight Training':
                        optional_attended += 1
                    elif session_type == 'Mission':
                        mission_attended += 1
                    elif session_type == 'Team Meeting':
                        team_meeting_attended += 1
                    elif session_type == 'Other':
                        other_attended += 1
            
            # DEBUG: print(f"DEBUG: PDF - Member attended {qualifying_attended} weekend training sessions in 6mo")
            
            # Create temp file
            fd, pdf_path = tempfile.mkstemp(suffix='.pdf')
            os.close(fd)
            
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            story.append(Paragraph(f"505 SAR Dogs Attendance Record for {member_name}", styles['Title']))
            story.append(Spacer(1, 12))
            
            # Summary lines (moved above table)
            summary_qualifying = f"You attended {qualifying_attended} out of {total_qualifying} weekend training sessions in the last 6 months."
            story.append(Paragraph(summary_qualifying, styles['Normal']))
            story.append(Spacer(1, 6))
            
            summary_optional = f"You attended {optional_attended} out of {total_optional} weeknight training sessions in the last 6 months."
            story.append(Paragraph(summary_optional, styles['Normal']))
            story.append(Spacer(1, 6))
            
            summary_mission = f"You attended {mission_attended} out of {total_mission} missions in the last 6 months."
            story.append(Paragraph(summary_mission, styles['Normal']))
            story.append(Spacer(1, 6))
            
            summary_team_meeting = f"You attended {team_meeting_attended} out of {total_team_meeting} team meetings in the last 6 months."
            story.append(Paragraph(summary_team_meeting, styles['Normal']))
            story.append(Spacer(1, 6))
            
            summary_other = f"You attended {other_attended} out of {total_other} other sessions in the last 6 months."
            story.append(Paragraph(summary_other, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Only build table if months_to_review > 0 and there are sessions
            if months_to_review > 0 and sessions_table:
                # Sort sessions chronologically by date (oldest first)
                sessions_sorted = sorted(sessions_table, key=lambda s: datetime.strptime(s.get('date', '01/01/2000'), "%m/%d/%Y"))
                
                # Build table data
                # NOTE: get_all_sessions renames session_date->date and session_type->type
                # First row is the header text spanning all columns
                month_text = "month" if months_to_review == 1 else "months"
                table_data = [[f'Recorded sessions over the last {months_to_review} {month_text}.', '', '', '']]
                # Second row is the column headers
                table_data.append(['Date', 'Location', 'Type', 'Attended'])
                
                for session in sessions_sorted:
                    session_id = session.get('id')
                    attended = self.db.get_attendance_status(session_id, member_id)
                    attended_str = "Yes" if attended else "No"
                    
                    # Convert old session type names to new names for display
                    # For 'Other' type, use the description instead
                    session_type = session.get('type', '')
                    if session_type == 'Qualifying Training':
                        session_type = 'Weekend Training'
                    elif session_type == 'Optional Training':
                        session_type = 'Weeknight Training'
                    elif session_type == 'Other':
                        description = session.get('description', '').strip()
                        if description:
                            session_type = description
                    
                    table_data.append([
                        session.get('date', ''),
                        session.get('location', ''),
                        session_type,
                        attended_str
                    ])
                
                # DEBUG: Print table data summary
                # print(f"DEBUG: PDF - Table has {len(table_data)} rows (including header)")
                # if len(table_data) > 2:
                #     print(f"DEBUG: PDF - First data row: {table_data[2]}")
                
                # Create table with repeatRows=2 to repeat header rows on page breaks
                table = Table(table_data, colWidths=[80, 150, 100, 60], repeatRows=2)
                table.setStyle(TableStyle([
                    # Spanning header row (row 0)
                    ('SPAN', (0, 0), (-1, 0)),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    # Column headers row (row 1)
                    ('BACKGROUND', (0, 1), (-1, 1), colors.grey),
                    ('TEXTCOLOR', (0, 1), (-1, 1), colors.whitesmoke),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 1), (-1, 1), 10),
                    ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
                    # Data rows (row 2+)
                    ('BACKGROUND', (0, 2), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                story.append(table)
            
            # Build PDF
            doc.build(story)
            return pdf_path
            
        except Exception as e:
            # DEBUG: print(f"DEBUG: PDF - Error generating PDF for {member.get('last_name', '')}: {e}")
            return None
        
    # ============================================================
    # Setup Tab
    # ============================================================
    
    def _build_setup_tab(self):
        """Build the Setup tab content."""
        # Main container
        main_frame = ttk.Frame(self.setup_tab, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Primary Storage Folder section
        self._build_primary_storage_section(main_frame)
        
        # Secondary Backup Folder section
        self._build_secondary_backup_section(main_frame)
        
        # Excel Participation File Location section
        self._build_excel_participation_section(main_frame)
        
        # Container for Email Configuration and PDF Presentation side by side
        email_pdf_frame = ttk.Frame(main_frame)
        email_pdf_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Email Configuration section (left side)
        self._build_email_config_section(email_pdf_frame)
        
        # PDF Presentation section (right side)
        # self._build_pdf_presentation_section(email_pdf_frame)
        
        # Default Values section
        self._build_default_values_section(main_frame)
        
    def _build_primary_storage_section(self, parent):
        """Build the Primary Storage Folder section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Primary Storage Folder", padding=5)
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = ttk.Frame(frame)
        row.pack(fill=tk.X, pady=5)
        
        # Path entry
        self.primary_path_entry = ttk.Entry(row, textvariable=self.vars.primary_storage_folder, width=50)
        self.primary_path_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Browse button
        ttk.Button(row, text="Browse", command=self._on_browse_primary).pack(side=tk.LEFT, padx=(0, 10))
        
        # Initialize Data Structures button
        ttk.Button(row, text="Initialize Data Structures", command=self._on_initialize_data).pack(side=tk.LEFT, padx=(0, 20))
        
    def _build_secondary_backup_section(self, parent):
        """Build the Secondary Backup Folder section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Secondary Backup Folder", padding=5)
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = ttk.Frame(frame)
        row.pack(fill=tk.X, pady=5)
        
        # Path entry
        self.secondary_path_entry = ttk.Entry(row, textvariable=self.vars.secondary_backup_folder, width=50)
        self.secondary_path_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Browse button
        ttk.Button(row, text="Browse", command=self._on_browse_secondary).pack(side=tk.LEFT, padx=(0, 10))
        
        # Restore from Secondary Backup button
        ttk.Button(row, text="Restore from Secondary Backup", command=self._on_restore_backup).pack(side=tk.LEFT)
        
    def _build_excel_participation_section(self, parent):
        """Build the Excel Participation File Location section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Excel Participation File Location (Folder)", padding=5)
        frame.pack(fill=tk.X, pady=(0, 10))
        
        row = ttk.Frame(frame)
        row.pack(fill=tk.X, pady=5)
        
        # Path entry
        self.excel_path_entry = ttk.Entry(row, textvariable=self.vars.excel_participation_folder, width=50)
        self.excel_path_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Browse button
        ttk.Button(row, text="Browse", command=self._on_browse_excel).pack(side=tk.LEFT)
        
    def _build_email_config_section(self, parent):
        """Build the Email Configuration section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Email Configuration (for monthly attendance reports)", padding=5)
        frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Row 1: SMTP Server, Port, Encryption
        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1, text="SMTP Server:", width=12).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.vars.smtp_server, width=20).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(row1, text="Port:", width=5).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.vars.smtp_port, width=6).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(row1, text="Encryption:", width=10).pack(side=tk.LEFT)
        encryption_combo = ttk.Combobox(row1, textvariable=self.vars.smtp_encryption, width=10, 
                                        values=["TLS", "STARTTLS", "SSL", "None"], state="readonly")
        encryption_combo.pack(side=tk.LEFT)
        
        # Row 2: Username and Sender Email
        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2, text="Username:", width=12).pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.vars.smtp_username, width=25).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(row2, text="Sender Email:", width=12).pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.vars.sender_email, width=25).pack(side=tk.LEFT)
        
        # Row 3: Password and Months to Review
        row3 = ttk.Frame(frame)
        row3.pack(fill=tk.X, pady=2)
        
        ttk.Label(row3, text="Password:", width=12).pack(side=tk.LEFT)
        self.password_entry = ttk.Entry(row3, textvariable=self.vars.sender_password, width=25, show="*")
        self.password_entry.pack(side=tk.LEFT)
        
        ttk.Label(row3, text="(App Password for Gmail)", font=('TkDefaultFont', 8)).pack(side=tk.LEFT, padx=(10, 20))
        
        # ttk.Label(row3, text="Months in PDF:").pack(side=tk.LEFT)
        # self.pdf_months_spinbox = ttk.Spinbox(
        #     row3, 
        #     textvariable=self.vars.pdf_months_to_review, 
        #     from_=0, 
        #     to=12, 
        #     width=6
        # )
        # self.pdf_months_spinbox.pack(side=tk.LEFT, padx=(5, 0))
        
    # def _build_pdf_presentation_section(self, parent):
    #     """Build the PDF Presentation section.
        
    #     Args:
    #         parent: Parent widget
    #     """
    #     frame = ttk.LabelFrame(parent, text="PDF Presentation", padding=5)
    #     frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(5, 0))
        
    #     # Review Attendance button
    #     ttk.Button(
    #         frame, 
    #         text="Review Attendance", 
    #         command=self._on_review_attendance
    #     ).pack(pady=2, padx=10)
        
    def _on_review_attendance(self):
        """Handle Review Attendance button click - show member selection dialog."""
        if not self.db.database_exists():
            messagebox.showwarning("Database Required", 
                                   "Please initialize the database first (Setup tab).")
            return
            
        members = self.db.get_all_members()
        if not members:
            messagebox.showinfo("No Members", "No members found in the database.")
            return
        
        # Show member selection dialog with date range
        self._show_attendance_review_dialog(members)
        
    def _build_default_values_section(self, parent):
        """Build the Default Values (Optional) section with Training Locations and Certifications.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Default Values", padding=5)
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        content = ttk.Frame(frame)
        content.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Left side: Training Locations
        self._build_training_locations_list(content)
        
        # Right side: Certifications
        self._build_certifications_list(content)
        
    def _build_training_locations_list(self, parent):
        """Build the Training Locations list section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Training Locations", padding=5)
        frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        content = ttk.Frame(frame)
        content.pack(fill=tk.BOTH, expand=True)
        
        # Left side: Listbox
        list_frame = ttk.Frame(content)
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Listbox with scrollbar
        listbox_container = ttk.Frame(list_frame)
        listbox_container.pack(fill=tk.BOTH, expand=True)
        
        self.locations_listbox = tk.Listbox(listbox_container, width=25, height=8, exportselection=False)
        self.locations_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(listbox_container, orient=tk.VERTICAL, command=self.locations_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.locations_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Bind selection event
        self.locations_listbox.bind('<<ListboxSelect>>', self._on_location_listbox_select)
        
        # Right side: Entry and buttons
        right_frame = ttk.Frame(content)
        right_frame.pack(side=tk.LEFT, anchor=tk.N)
        
        # Location entry
        ttk.Label(right_frame, text="Location:").pack(anchor=tk.W)
        
        self.new_location_var = tk.StringVar()
        self.location_entry = ttk.Entry(right_frame, textvariable=self.new_location_var, width=20)
        self.location_entry.pack(anchor=tk.W, pady=(0, 5))
        self.location_entry.bind('<Return>', self._on_add_location_enter)
        
        # Add Location button
        ttk.Button(right_frame, text="Add Location", command=self._on_add_location).pack(anchor=tk.W, pady=(0, 5), fill=tk.X)
        
        # Remove Selected button
        self.remove_location_btn = ttk.Button(right_frame, text="Remove Selected", 
                                               command=self._on_remove_location, state=tk.DISABLED)
        self.remove_location_btn.pack(anchor=tk.W, fill=tk.X)
        
    def _build_certifications_list(self, parent):
        """Build the Certifications list section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Certifications", padding=5)
        frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        content = ttk.Frame(frame)
        content.pack(fill=tk.BOTH, expand=True)
        
        # Left side: Listbox
        list_frame = ttk.Frame(content)
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # Listbox with scrollbar
        listbox_container = ttk.Frame(list_frame)
        listbox_container.pack(fill=tk.BOTH, expand=True)
        
        self.certifications_listbox = tk.Listbox(listbox_container, width=25, height=8, exportselection=False)
        self.certifications_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(listbox_container, orient=tk.VERTICAL, command=self.certifications_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.certifications_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Bind selection event
        self.certifications_listbox.bind('<<ListboxSelect>>', self._on_certification_listbox_select)
        
        # Right side: Entry and buttons
        right_frame = ttk.Frame(content)
        right_frame.pack(side=tk.LEFT, anchor=tk.N)
        
        # Certification entry
        ttk.Label(right_frame, text="Certification:").pack(anchor=tk.W)
        
        self.new_certification_var = tk.StringVar()
        self.certification_entry = ttk.Entry(right_frame, textvariable=self.new_certification_var, width=20)
        self.certification_entry.pack(anchor=tk.W, pady=(0, 5))
        self.certification_entry.bind('<Return>', self._on_add_certification_enter)
        
        # Add Certification button
        ttk.Button(right_frame, text="Add Certification", command=self._on_add_certification).pack(anchor=tk.W, pady=(0, 5), fill=tk.X)
        
        # Remove Selected button
        self.remove_certification_btn = ttk.Button(right_frame, text="Remove Selected", 
                                                    command=self._on_remove_certification, state=tk.DISABLED)
        self.remove_certification_btn.pack(anchor=tk.W, fill=tk.X)
        
    # ============================================================
    # Event Handlers - Setup Tab
    # ============================================================
    
    def _on_browse_primary(self):
        """Handle Browse button click for primary storage folder."""
        from tkinter import filedialog
        folder = filedialog.askdirectory(
            title="Select Primary Storage Folder",
            initialdir=self.vars.primary_storage_folder.get() or "."
        )
        if folder:
            self.vars.primary_storage_folder.set(folder)
            
    def _on_browse_secondary(self):
        """Handle Browse button click for secondary backup folder."""
        from tkinter import filedialog
        folder = filedialog.askdirectory(
            title="Select Secondary Backup Folder",
            initialdir=self.vars.secondary_backup_folder.get() or "."
        )
        if folder:
            self.vars.secondary_backup_folder.set(folder)
            
    def _on_browse_excel(self):
        """Handle Browse button click for Excel participation folder."""
        from tkinter import filedialog
        folder = filedialog.askdirectory(
            title="Select Excel Participation File Location",
            initialdir=self.vars.excel_participation_folder.get() or "."
        )
        if folder:
            self.vars.excel_participation_folder.set(folder)
            
    def _on_initialize_data(self):
        """Handle Initialize Data Structures button click."""
        primary_path = self.vars.primary_storage_folder.get()
        
        if not primary_path:
            messagebox.showwarning("Initialize Data", 
                                   "Please select a Primary Storage Folder first.")
            return
            
        # Set database path
        db_path = os.path.join(primary_path, "training_tracker.db")
        self.db.set_db_path(db_path)
        
        # Check if database already exists
        replace_existing = False
        if self.db.database_exists():
            if not messagebox.askyesno("Database Exists",
                                       f"A database already exists at:\n{db_path}\n\n"
                                       "Do you want to replace it?\n"
                                       "WARNING: All existing data will be lost!"):
                return
            replace_existing = True
        else:
            if not messagebox.askyesno("Initialize Data Structures",
                                       f"This will create the database at:\n{db_path}\n\n"
                                       "Continue?"):
                return
        
        # Initialize the database
        if self.db.initialize_database(replace_existing=replace_existing):
            # Save config with the new path
            self.config.primary_storage_folder = primary_path
            self.config.save()
            
            self._load_data_from_database()
        else:
            messagebox.showerror("Error", "Failed to initialize database.")
            
    def _load_data_from_database(self):
        """Load all data from the database and refresh UI."""
        if not self.db.database_exists():
            return
            
        try:
            # Load training locations
            self._refresh_locations_listbox()
            
            # Load certification types
            self._refresh_certifications_listbox()
            self._refresh_cert_tree()
            
            # Load members
            members = self.db.get_all_members()
            self.ui_state.members_list = members
            self._refresh_members_tree()
            self.refresh_last_name_combobox()
            
            # Load sessions
            sessions = self.db.get_all_sessions()
            self.ui_state.sessions_list = sessions
            self.refresh_location_combobox()
            
            # Refresh attendance tree to show all members
            self._refresh_attendance_tree()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")
        
    def _on_restore_backup(self):
        """Handle Restore from Secondary Backup button click."""
        secondary_path = self.vars.secondary_backup_folder.get()
        
        if not secondary_path:
            messagebox.showwarning("Restore Backup", 
                                   "Please select a Secondary Backup Folder first.")
            return
            
        # Get list of backups
        backups = config.get_backup_files(secondary_path)
        
        if not backups:
            messagebox.showinfo("Restore Backup", "No backup files found in the backup folder.")
            return
            
        # Show backup selection popup
        self._show_backup_selection_popup(backups)
        
    def _show_backup_selection_popup(self, backups: list):
        """Show a popup window to select a backup to restore.
        
        Args:
            backups: List of (filepath, timestamp) tuples
        """
        popup = tk.Toplevel(self.root)
        popup.title("Select Backup to Restore")
        popup.geometry("500x350")
        popup.transient(self.root)
        popup.grab_set()
        
        # Center on parent
        popup.geometry(f"+{self.root.winfo_x() + 100}+{self.root.winfo_y() + 100}")
        
        # Instructions
        ttk.Label(popup, text="Select a backup to restore (newest first):").pack(pady=(10, 5), padx=10, anchor=tk.W)
        
        # Listbox with scrollbar
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, height=12)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        backup_map = {}
        for filepath, timestamp in backups:
            display_text = f"{timestamp.strftime('%Y-%m-%d %H:%M:%S')} - {os.path.basename(filepath)}"
            listbox.insert(tk.END, display_text)
            backup_map[display_text] = filepath
            
        # Button frame
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def on_restore():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Select Backup", "Please select a backup to restore.")
                return
                
            display_text = listbox.get(selection[0])
            filepath = backup_map.get(display_text)
            
            if not filepath:
                return
                
            if messagebox.askyesno("Confirm Restore",
                                   "This will replace all current data with the backup.\n\n"
                                   "Are you sure you want to continue?"):
                popup.destroy()
                self._restore_from_backup_file(filepath)
                
        def on_cancel():
            popup.destroy()
            
        ttk.Button(btn_frame, text="Restore Selected", command=on_restore).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.LEFT, padx=5)
        
    def _restore_from_backup_file(self, filepath: str):
        """Restore data from a backup file.
        
        Args:
            filepath: Path to the backup file
        """
        # Load backup data
        data = config.load_backup(filepath)
        if not data:
            messagebox.showerror("Restore Error", "Failed to load backup file.")
            return
            
        # Re-initialize database
        primary_path = self.vars.primary_storage_folder.get()
        if not primary_path:
            messagebox.showerror("Restore Error", "Primary storage folder not set.")
            return
            
        db_path = os.path.join(primary_path, "training_tracker.db")
        self.db.set_db_path(db_path)
        
        # Initialize fresh database
        if not self.db.initialize_database(replace_existing=True):
            messagebox.showerror("Restore Error", "Failed to initialize database.")
            return
            
        # Import data
        success, message = self.db.import_all_data(data)
        
        if success:
            self._load_data_from_database()
        else:
            messagebox.showerror("Restore Error", f"Failed to restore data:\n{message}")
        
    def refresh_user_combobox(self, users: list):
        """Refresh the user combobox with available users.
        
        Args:
            users: List of user names
        """
        # User combobox removed - this method kept for compatibility
        pass
        
    def _on_location_listbox_select(self, event=None):
        """Handle selection change in locations listbox.
        
        Args:
            event: The tkinter event
        """
        selection = self.locations_listbox.curselection()
        if selection:
            self.remove_location_btn.configure(state=tk.NORMAL)
        else:
            self.remove_location_btn.configure(state=tk.DISABLED)
            
    def _on_add_location_enter(self, event=None):
        """Handle Enter key in location entry.
        
        Args:
            event: The tkinter event
        """
        self._on_add_location()
        
    def _on_add_location(self):
        """Handle Add Location button click."""
        location = self.new_location_var.get().strip().title()
        
        if not location:
            messagebox.showwarning("Add Location", "Please enter a location name.")
            return
        
        # Update the entry field to show title case
        self.new_location_var.set(location)
            
        if not self.db.database_exists():
            messagebox.showwarning("Add Location", 
                                   "Please initialize the database first.")
            return
            
        if self.db.add_training_location(location):
            self.new_location_var.set("")
            self._refresh_locations_listbox()
            self.refresh_location_combobox()
        else:
            messagebox.showwarning("Add Location", 
                                   f"Location '{location}' already exists or could not be added.")
            
    def _on_remove_location(self):
        """Handle Remove Selected button click for locations."""
        selection = self.locations_listbox.curselection()
        if not selection:
            return
            
        location = self.locations_listbox.get(selection[0])
        
        if not messagebox.askyesno("Remove Location",
                                   f"Remove location '{location}'?"):
            return
            
        if self.db.remove_training_location(location):
            self._refresh_locations_listbox()
            self.refresh_location_combobox()
        else:
            messagebox.showerror("Error", "Failed to remove location.")
            
    def _refresh_locations_listbox(self):
        """Refresh the locations listbox from database."""
        self.locations_listbox.delete(0, tk.END)
        
        if self.db.database_exists():
            locations = self.db.get_training_locations()
            for loc in locations:
                self.locations_listbox.insert(tk.END, loc)
                
        self.remove_location_btn.configure(state=tk.DISABLED)
        
    def _on_certification_listbox_select(self, event=None):
        """Handle selection change in certifications listbox.
        
        Args:
            event: The tkinter event
        """
        selection = self.certifications_listbox.curselection()
        if selection:
            self.remove_certification_btn.configure(state=tk.NORMAL)
        else:
            self.remove_certification_btn.configure(state=tk.DISABLED)
            
    def _on_add_certification_enter(self, event=None):
        """Handle Enter key in certification entry.
        
        Args:
            event: The tkinter event
        """
        self._on_add_certification()
        
    def _on_add_certification(self):
        """Handle Add Certification button click."""
        cert_name = self.new_certification_var.get().strip()
        
        if not cert_name:
            messagebox.showwarning("Add Certification", "Please enter a certification name.")
            return
            
        if not self.db.database_exists():
            messagebox.showwarning("Add Certification", 
                                   "Please initialize the database first.")
            return
            
        if self.db.add_certification_type(cert_name):
            self.new_certification_var.set("")
            self._refresh_certifications_listbox()
            self._refresh_cert_tree()  # Update demographics tab
        else:
            messagebox.showwarning("Add Certification", 
                                   f"Certification '{cert_name}' already exists or could not be added.")
            
    def _on_remove_certification(self):
        """Handle Remove Selected button click for certifications."""
        selection = self.certifications_listbox.curselection()
        if not selection:
            return
            
        cert_name = self.certifications_listbox.get(selection[0])
        
        if not messagebox.askyesno("Remove Certification",
                                   f"Remove certification '{cert_name}'?\n\n"
                                   "This will not delete existing certification dates for members."):
            return
            
        if self.db.remove_certification_type(cert_name):
            self._refresh_certifications_listbox()
            self._refresh_cert_tree()  # Update demographics tab
        else:
            messagebox.showerror("Error", "Failed to remove certification.")
            
    def _refresh_certifications_listbox(self):
        """Refresh the certifications listbox from database."""
        self.certifications_listbox.delete(0, tk.END)
        
        if self.db.database_exists():
            certs = self.db.get_certification_types()
            for cert in certs:
                self.certifications_listbox.insert(tk.END, cert)
                
        self.remove_certification_btn.configure(state=tk.DISABLED)
        
    # ============================================================
    # Demographics Tab
    # ============================================================
    
    def _build_demographics_tab(self):
        """Build the Demographics tab content."""
        # Main container with scrollable frame
        main_container = ttk.Frame(self.demographics_tab)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Top frame for form fields
        form_frame = ttk.LabelFrame(main_container, text="Member Information", padding=10)
        form_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Build form sections
        self._build_personal_info_section(form_frame)
        self._build_contact_info_section(form_frame)
        self._build_emergency_contact_section(form_frame)
        # Note: Ham Callsign and Mission Eligible are now in _build_personal_info_section
        
        # Members list treeview (moved above certifications)
        self._build_members_list_section(main_container)
        
        # Certifications and buttons in a horizontal frame
        cert_buttons_frame = ttk.Frame(main_container)
        cert_buttons_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Certifications treeview on the left
        self._build_certifications_section(cert_buttons_frame)
        
        # Buttons frame on the right of certifications
        self._build_demographics_buttons(cert_buttons_frame)
        
        # Set up name field change tracking for duplicate detection
        self.vars.first_name.trace_add("write", self._on_name_field_changed)
        self.vars.last_name.trace_add("write", self._on_name_field_changed)
        
        # Initial state - disable fields until name is entered
        self._update_demographics_entry_states()
        
    def _build_personal_info_section(self, parent):
        """Build the personal information section.
        
        Args:
            parent: Parent widget
        """
        # Initialize list to store entry widgets that should be disabled when name is empty
        self.demographics_entry_widgets = []
        
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=5)
        
        # Row 1: Last Name (combobox), First Name, Ham Callsign
        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1, text="Last Name:", width=15).pack(side=tk.LEFT)
        self.last_name_combo = ttk.Combobox(row1, textvariable=self.vars.last_name, width=23)
        self.last_name_combo.pack(side=tk.LEFT, padx=(0, 20))
        self.last_name_combo.bind("<<ComboboxSelected>>", self._on_last_name_selected)
        
        ttk.Label(row1, text="First Name:", width=15).pack(side=tk.LEFT)
        first_name_entry = tk.Entry(row1, textvariable=self.vars.first_name, width=25)
        first_name_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(row1, text="Ham Callsign:", width=12).pack(side=tk.LEFT)
        self.callsign_entry = tk.Entry(row1, textvariable=self.vars.ham_callsign, width=15)
        self.callsign_entry.pack(side=tk.LEFT)
        self.demographics_entry_widgets.append(self.callsign_entry)
        
        # Row 2: Address
        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2, text="Address:", width=15).pack(side=tk.LEFT)
        self.address_entry = tk.Entry(row2, textvariable=self.vars.address, width=60)
        self.address_entry.pack(side=tk.LEFT)
        self.demographics_entry_widgets.append(self.address_entry)
        
        # Row 3: Status radiobuttons (Member/Candidate/Guest) and Mission Eligible checkbox
        row3 = ttk.Frame(frame)
        row3.pack(fill=tk.X, pady=2)
        
        ttk.Label(row3, text="Status:", width=15).pack(side=tk.LEFT)
        
        # Status radiobuttons
        self.status_radios = []
        for status in ui_support.get_member_statuses():
            rb = ttk.Radiobutton(
                row3,
                text=status,
                variable=self.vars.member_status,
                value=status,
                command=self._on_member_status_changed
            )
            rb.pack(side=tk.LEFT, padx=5)
            self.status_radios.append(rb)
            self.demographics_entry_widgets.append(rb)
        
        # Mission Eligible checkbox (to the right of status radiobuttons)
        self.mission_check = ttk.Checkbutton(
            row3, 
            text="Mission Eligible",
            variable=self.vars.mission_eligible
        )
        self.mission_check.pack(side=tk.LEFT, padx=20)
        self.demographics_entry_widgets.append(self.mission_check)
        
    def _build_contact_info_section(self, parent):
        """Build the contact information section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Contact Information", padding=5)
        frame.pack(fill=tk.X, pady=5)
        
        # Row 1: Cell Phone, Email (swapped - Email moved up from Row 2)
        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1, text="Cell Phone:", width=15).pack(side=tk.LEFT)
        self.cell_entry = tk.Entry(row1, textvariable=self.vars.cell_phone, width=20)
        self.cell_entry.pack(side=tk.LEFT, padx=(0, 20))
        self.cell_entry.bind('<FocusOut>', self._on_phone_focus_out)
        self.demographics_entry_widgets.append(self.cell_entry)
        
        ttk.Label(row1, text="Email:", width=15).pack(side=tk.LEFT)
        self.email_entry = tk.Entry(row1, textvariable=self.vars.email, width=30)
        self.email_entry.pack(side=tk.LEFT)
        self.demographics_entry_widgets.append(self.email_entry)
        
        # Row 2: Home Phone, Alternate Email (swapped - Home Phone moved down from Row 1)
        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2, text="Home Phone:", width=15).pack(side=tk.LEFT)
        self.home_entry = tk.Entry(row2, textvariable=self.vars.home_phone, width=20)
        self.home_entry.pack(side=tk.LEFT, padx=(0, 20))
        self.home_entry.bind('<FocusOut>', self._on_phone_focus_out)
        self.demographics_entry_widgets.append(self.home_entry)
        
        ttk.Label(row2, text="Alternate Email:", width=15).pack(side=tk.LEFT)
        self.alt_email_entry = tk.Entry(row2, textvariable=self.vars.alternate_email, width=30)
        self.alt_email_entry.pack(side=tk.LEFT)
        self.demographics_entry_widgets.append(self.alt_email_entry)
        
    def _build_emergency_contact_section(self, parent):
        """Build the emergency contact section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Emergency Contact", padding=5)
        frame.pack(fill=tk.X, pady=5)
        
        row = ttk.Frame(frame)
        row.pack(fill=tk.X, pady=2)
        
        ttk.Label(row, text="Contact Name:", width=15).pack(side=tk.LEFT)
        self.emergency_name_entry = tk.Entry(row, textvariable=self.vars.emergency_contact_name, width=25)
        self.emergency_name_entry.pack(side=tk.LEFT, padx=(0, 20))
        self.demographics_entry_widgets.append(self.emergency_name_entry)
        
        ttk.Label(row, text="Phone:", width=10).pack(side=tk.LEFT)
        self.emergency_phone_entry = tk.Entry(row, textvariable=self.vars.emergency_contact_phone, width=20)
        self.emergency_phone_entry.pack(side=tk.LEFT)
        self.emergency_phone_entry.bind('<FocusOut>', self._on_phone_focus_out)
        self.demographics_entry_widgets.append(self.emergency_phone_entry)
        
    def _build_additional_info_section(self, parent):
        """Build the additional information section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(frame, text="Ham Callsign:", width=15).pack(side=tk.LEFT)
        self.callsign_entry = tk.Entry(frame, textvariable=self.vars.ham_callsign, width=15)
        self.callsign_entry.pack(side=tk.LEFT, padx=(0, 20))
        self.demographics_entry_widgets.append(self.callsign_entry)
        
        self.mission_check = ttk.Checkbutton(
            frame, 
            text="Mission Eligible",
            variable=self.vars.mission_eligible
        )
        self.mission_check.pack(side=tk.LEFT, padx=20)
        self.demographics_entry_widgets.append(self.mission_check)
        
    def _build_certifications_section(self, parent):
        """Build the certifications treeview section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Certifications (double-click to edit)", padding=5)
        frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Create treeview with columns
        columns = ("certification", "date")
        self.cert_tree = ttk.Treeview(frame, columns=columns, show="headings", height=6)
        
        # Configure columns - certification column narrower
        self.cert_tree.heading("certification", text="Certification")
        self.cert_tree.heading("date", text="Date Completed")
        
        self.cert_tree.column("certification", width=150, anchor=tk.W)
        self.cert_tree.column("date", width=100, anchor=tk.CENTER)
        
        # Configure alternating row colors
        self.cert_tree.tag_configure('oddrow', background='#f0f0f0')
        self.cert_tree.tag_configure('evenrow', background='white')
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.cert_tree.yview)
        self.cert_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        self.cert_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind double-click to edit date
        self.cert_tree.bind("<Double-1>", self._on_cert_tree_double_click)
        
    def _refresh_cert_tree(self):
        """Refresh the certification treeview with current certification types."""
        # Clear current items
        for item in self.cert_tree.get_children():
            self.cert_tree.delete(item)
            
        # Get certification types from database or use defaults
        if self.db.database_exists():
            cert_names = self.db.get_certification_types()
        else:
            cert_names = ui_support.get_certification_row_names()
            
        # Add rows with alternating colors
        for i, name in enumerate(cert_names):
            tag = 'oddrow' if i % 2 else 'evenrow'
            self.cert_tree.insert("", tk.END, values=(name, ""), iid=f"cert_{i}", tags=(tag,))
            
        # Update from current vars if editing a member
        self._update_cert_tree_from_vars()
        
    def _on_cert_tree_double_click(self, event):
        """Handle double-click on certification treeview to edit date.
        
        Args:
            event: The tkinter event
        """
        # Get the item and column that was clicked
        region = self.cert_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
            
        column = self.cert_tree.identify_column(event.x)
        item = self.cert_tree.identify_row(event.y)
        
        if not item:
            return
            
        # Get current values
        values = self.cert_tree.item(item, "values")
        cert_name = values[0]
        current_date = values[1] if len(values) > 1 else ""
        
        # Show date picker popup
        self._show_cert_date_popup(item, cert_name, current_date)
        
    def _show_cert_date_popup(self, item_id: str, cert_name: str, current_date: str):
        """Show a popup to edit certification date.
        
        Args:
            item_id: Treeview item ID
            cert_name: Name of the certification
            current_date: Current date value (may be empty)
        """
        popup = tk.Toplevel(self.root)
        popup.title(f"Edit: {cert_name}")
        popup.geometry("300x150")
        popup.transient(self.root)
        popup.grab_set()
        
        # Center on parent
        popup.geometry(f"+{self.root.winfo_x() + 200}+{self.root.winfo_y() + 200}")
        
        # Label
        ttk.Label(popup, text=f"Date Completed for:\n{cert_name}").pack(pady=(10, 5))
        
        # Date picker frame
        date_frame = ttk.Frame(popup)
        date_frame.pack(pady=10)
        
        # Use DateEntry if available, otherwise regular entry
        if ui_support.TKCALENDAR_AVAILABLE:
            from tkcalendar import DateEntry
            date_var = tk.StringVar(value=current_date)
            date_entry = DateEntry(date_frame, textvariable=date_var, date_pattern='mm/dd/yyyy', width=12)
            
            # Set the date if current_date is valid
            if current_date:
                try:
                    parsed_date = datetime.strptime(current_date, "%m/%d/%Y")
                    date_entry.set_date(parsed_date)
                except ValueError:
                    pass
            date_entry.pack()
        else:
            date_var = tk.StringVar(value=current_date)
            date_entry = ttk.Entry(date_frame, textvariable=date_var, width=15)
            date_entry.pack()
            ttk.Label(date_frame, text="(MM/DD/YYYY)").pack()
        
        # Button frame
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=10)
        
        def on_save():
            new_date = date_var.get().strip()
            
            # Validate date if provided
            if new_date:
                is_valid, result = ui_support.validate_date(new_date)
                if not is_valid:
                    messagebox.showwarning("Invalid Date", result)
                    return
                new_date = result
                
            # Update treeview
            values = self.cert_tree.item(item_id, "values")
            self.cert_tree.item(item_id, values=(values[0], new_date))
            
            # Auto-save if member is selected
            self._auto_save_certification(values[0], new_date)
            
            popup.destroy()
            
        def on_delete():
            # Clear the date
            values = self.cert_tree.item(item_id, "values")
            self.cert_tree.item(item_id, values=(values[0], ""))
            
            # Auto-save the empty date
            self._auto_save_certification(values[0], "")
            
            popup.destroy()
        
        ttk.Button(btn_frame, text="Save", command=on_save).pack(side=tk.LEFT, padx=5)
        
        # Show Delete button if there's an existing date, otherwise Cancel
        if current_date:
            ttk.Button(btn_frame, text="Delete", command=on_delete).pack(side=tk.LEFT, padx=5)
        else:
            ttk.Button(btn_frame, text="Cancel", command=popup.destroy).pack(side=tk.LEFT, padx=5)
            
    def _auto_save_certification(self, cert_name: str, date_value: str):
        """Auto-save a certification date if a member is selected.
        
        Args:
            cert_name: Name of the certification
            date_value: Date value to save
        """
        member_id = self.vars.selected_member_id.get()
        
        # Update the StringVar for this certification
        cert_var_map = {
            "Pack Check": self.vars.pack_check_date,
            "On-line Base Medical": self.vars.online_base_medical_date,
            "Crime Scene Preservation": self.vars.crime_scene_preservation_date,
            "Blood-borne Pathogens": self.vars.blood_borne_pathogens_date,
            "NM SAR Field Certification": self.vars.nm_sar_field_certification_date,
            "Fitness Hike 1": self.vars.fitness_hike_1_date,
            "Fitness Hike 2": self.vars.fitness_hike_2_date,
            "Fitness Hike 3": self.vars.fitness_hike_3_date,
        }
        
        if cert_name in cert_var_map:
            cert_var_map[cert_name].set(date_value)
        
        if member_id > 0 and self.db.database_exists():
            # Update the certification in the database
            self.db._save_member_certifications(member_id, {cert_name: date_value})
        
    def _build_demographics_buttons(self, parent):
        """Build the demographics action buttons.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.Frame(parent)
        frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        ttk.Button(frame, text="New Member", command=self._on_new_member, width=15).pack(pady=2)
        ttk.Button(frame, text="Save Member", command=self._on_save_member, width=15).pack(pady=2)
        ttk.Button(frame, text="Delete Member", command=self._on_delete_member, width=15).pack(pady=2)
        ttk.Button(frame, text="Clear Form", command=self._on_clear_demographics, width=15).pack(pady=2)
        
        
    def _build_members_list_section(self, parent):
        """Build the members list treeview section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Members List", padding=5)
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Create treeview with columns
        columns = ("name", "callsign", "eligible", "cell_phone")
        self.members_tree = ttk.Treeview(frame, columns=columns, show="headings", height=6)
        
        # Configure columns
        self.members_tree.heading("name", text="Name")
        self.members_tree.heading("callsign", text="Callsign")
        self.members_tree.heading("eligible", text="Mission Eligible")
        self.members_tree.heading("cell_phone", text="Cell Phone")
        
        self.members_tree.column("name", width=200, anchor=tk.W)
        self.members_tree.column("callsign", width=100, anchor=tk.CENTER)
        self.members_tree.column("eligible", width=120, anchor=tk.CENTER)
        self.members_tree.column("cell_phone", width=150, anchor=tk.CENTER)
        
        # Configure alternating row colors
        self.members_tree.tag_configure('oddrow', background='#f0f0f0')
        self.members_tree.tag_configure('evenrow', background='white')
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.members_tree.yview)
        self.members_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        self.members_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind selection event
        self.members_tree.bind("<<TreeviewSelect>>", self._on_member_selected)
        
    # ============================================================
    # Training Sessions Tab
    # ============================================================
    
    def _build_training_sessions_tab(self):
        """Build the Training Sessions tab content."""
        main_container = ttk.Frame(self.training_sessions_tab)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Session details frame
        self._build_session_details_section(main_container)
        
        # Session buttons
        self._build_session_buttons(main_container)
        
        # Attendance treeview
        self._build_attendance_section(main_container)
        
    def _build_session_details_section(self, parent):
        """Build the session details section.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.LabelFrame(parent, text="Session Details", padding=10)
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Top row: Date, Location, Session Type
        row = ttk.Frame(frame)
        row.pack(fill=tk.X, pady=5)
        
        # Column 1: Date (using custom datepicker)
        date_frame = ttk.Frame(row)
        date_frame.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(date_frame, text="Date:").pack(anchor=tk.W)
        
        # Use DateEntry if available, otherwise regular entry
        if ui_support.TKCALENDAR_AVAILABLE:
            self.date_entry = ui_support.TrainingDateEntry(
                date_frame,
                sessions_callback=self._get_sessions_for_calendar,
                textvariable=self.vars.session_date,
                width=12,
                date_pattern='mm/dd/yyyy'
            )
            self.date_entry.bind("<<DateEntrySelected>>", self._on_session_date_selected)
            # Clear the default date - leave blank for new entries
            self.date_entry.delete(0, tk.END)
            self.vars.session_date.set("")
        else:
            self.date_entry = ttk.Entry(date_frame, textvariable=self.vars.session_date, width=15)
            self.date_entry.bind('<FocusOut>', self._on_session_date_focus_out)
        self.date_entry.pack(anchor=tk.W)
        
        # Column 2: Location (combobox)
        loc_frame = ttk.Frame(row)
        loc_frame.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Label(loc_frame, text="Location:").pack(anchor=tk.W)
        self.location_combo = ttk.Combobox(loc_frame, textvariable=self.vars.session_location, width=28)
        self.location_combo.pack(anchor=tk.W)
        
        # Column 3: Session Type (Radio buttons)
        type_frame = ttk.LabelFrame(row, text="Session Type:", padding=5)
        type_frame.pack(side=tk.LEFT, padx=(0, 5))
        
        self.session_type_radios = []
        for session_type in ui_support.get_session_types():
            rb = ttk.Radiobutton(
                type_frame,
                text=session_type,
                variable=self.vars.session_type,
                value=session_type,
                command=self._on_session_type_changed
            )
            rb.pack(side=tk.LEFT, padx=5)
            self.session_type_radios.append(rb)
            
        # Description row (initially hidden, shown for Mission/Other)
        self.description_frame = ttk.Frame(frame)
        # Don't pack yet - will be shown/hidden based on session type
        
        ttk.Label(self.description_frame, text="Description:", width=12).pack(side=tk.LEFT)
        description_entry = ttk.Entry(self.description_frame, textvariable=self.vars.session_description, width=50)
        description_entry.pack(side=tk.LEFT, padx=(0, 10))
            
    def _build_session_buttons(self, parent):
        """Build the session action buttons.
        
        Args:
            parent: Parent widget
        """
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(frame, text="New Session", command=self._on_new_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Delete Session", command=self._on_delete_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Clear Form", command=self._on_clear_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame, text="Modify Date and/or Location", command=self._on_modify_session).pack(side=tk.LEFT, padx=5)

        # Spinbox for months to email on exit
        self.pdf_months_spinbox = ttk.Spinbox(
            frame, 
            textvariable=self.vars.pdf_months_to_review, 
            from_=0, 
            to=12, 
            width=4
        )
        self.pdf_months_spinbox.pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Label(frame, text="Months to Review in email PDF:").pack(side=tk.RIGHT)

        # Review Attendance button
        ttk.Button(
            frame, 
            text="Review Attendance", 
            command=self._on_review_attendance
        ).pack(side=tk.RIGHT,pady=2, padx=10)
        
        
    def _build_attendance_section(self, parent):
        """Build the attendance tracking treeview section.
        
        Args:
            parent: Parent widget
        """
        self.attendance_frame = ttk.LabelFrame(parent, text="Member Attendance", padding=5)
        self.attendance_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Top row with instructions and sort options
        top_row = ttk.Frame(self.attendance_frame)
        top_row.pack(fill=tk.X, pady=(0, 5))
        
        # Instructions (left side)
        self.attendance_instructions = ttk.Label(top_row, text="Double-click 'Attended' column to toggle Yes/No")
        self.attendance_instructions.pack(side=tk.LEFT, anchor=tk.W)
        
        # Sort options (right side)
        sort_frame = ttk.Frame(top_row)
        sort_frame.pack(side=tk.RIGHT)
        
        ttk.Label(sort_frame, text="Sort by:").pack(side=tk.LEFT, padx=(0, 5))
        
        # Variable for sort order - default to last name
        self.attendance_sort_var = tk.StringVar(value="last")
        
        ttk.Radiobutton(sort_frame, text="Last Name", variable=self.attendance_sort_var, 
                        value="last", command=self._on_attendance_sort_changed).pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(sort_frame, text="First Name", variable=self.attendance_sort_var, 
                        value="first", command=self._on_attendance_sort_changed).pack(side=tk.LEFT, padx=2)
        
        # Create treeview with columns (Status added, weekend_count commented out)
        columns = ("name", "status", "attended")
        # COMMENTED OUT: columns = ("name", "status", "attended", "weekend_count")
        self.attendance_tree = ttk.Treeview(self.attendance_frame, columns=columns, show="headings", height=12)
        
        # Configure columns
        self.attendance_tree.heading("name", text="Name")
        self.attendance_tree.heading("status", text="Status")
        self.attendance_tree.heading("attended", text="Attended")
        # COMMENTED OUT: self.attendance_tree.heading("weekend_count", text="Weekend Training Last 6 mo")
        
        self.attendance_tree.column("name", width=200, anchor=tk.W)
        self.attendance_tree.column("status", width=80, anchor=tk.CENTER)
        self.attendance_tree.column("attended", width=100, anchor=tk.CENTER)
        # COMMENTED OUT: self.attendance_tree.column("weekend_count", width=200, anchor=tk.CENTER)
        
        # Configure alternating row colors
        self.attendance_tree.tag_configure('oddrow', background='#f0f0f0')
        self.attendance_tree.tag_configure('evenrow', background='white')
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(self.attendance_frame, orient=tk.VERTICAL, command=self.attendance_tree.yview)
        self.attendance_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        self.attendance_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind double-click to toggle attendance
        self.attendance_tree.bind("<Double-1>", self._on_attendance_double_click)
        
        # Initial state - disabled until date is selected
        self._update_attendance_section_state()
        
    def _on_attendance_sort_changed(self):
        """Handle change in attendance list sort order."""
        self._refresh_attendance_tree()
        
    # ============================================================
    # Event Handlers - Demographics Tab
    # ============================================================
    
    def _on_phone_focus_out(self, event):
        """Handle phone field focus out for formatting.
        
        Args:
            event: The tkinter event
        """
        widget = event.widget
        # Get the StringVar associated with this entry
        for var in [self.vars.cell_phone, self.vars.home_phone, self.vars.emergency_contact_phone]:
            try:
                if str(widget.cget('textvariable')) == str(var):
                    phone = var.get()
                    if phone:
                        is_valid, result = ui_support.validate_phone(phone)
                        if is_valid:
                            var.set(result)
                    break
            except:
                pass
    
    def _on_name_field_changed(self, *args):
        """Handle changes to first or last name fields."""
        self._update_demographics_entry_states()
        
        # Only check for duplicates if user is typing (not loading from selection)
        # and not currently editing an existing member
        if not self.ui_state.is_editing_member and not self._loading_member_data:
            self._check_duplicate_name()
            
    def _check_duplicate_name(self):
        """Check if the entered name already exists in the database."""
        first_name = self.vars.first_name.get().strip()
        last_name = self.vars.last_name.get().strip()
        
        if not first_name or not last_name:
            return
            
        if not self.db.database_exists():
            return
            
        # Check if member with this name already exists
        existing = self.db.get_member_by_name(first_name, last_name)
        
        if existing and existing.get('id') != self.vars.selected_member_id.get():
            response = messagebox.askyesno(
                "Duplicate Name",
                f"Individual with the name '{first_name} {last_name}' already exists.\n\n"
                "Would you like to load and update their information?"
            )
            
            if response:
                self._load_member_data(existing)
                
    def _update_demographics_entry_states(self):
        """Enable/disable demographics entry fields based on name fields.
        
        All fields except name fields and members list are disabled when
        either first or last name is blank.
        """
        first_name = self.vars.first_name.get().strip()
        last_name = self.vars.last_name.get().strip()
        
        # Determine if fields should be enabled - BOTH names must have values
        enable = bool(first_name and last_name)
        state = tk.NORMAL if enable else tk.DISABLED
        
        # Update all entry widgets in demographics (except name fields)
        for widget in self.demographics_entry_widgets:
            try:
                widget.configure(state=state)
            except tk.TclError:
                pass
                
        # Also disable/enable the certifications treeview
        if self.cert_tree:
            try:
                if enable:
                    self.cert_tree.bind("<Double-1>", self._on_cert_tree_double_click)
                    self.cert_tree.configure(selectmode='browse')
                else:
                    self.cert_tree.unbind("<Double-1>")
                    self.cert_tree.configure(selectmode='none')
            except tk.TclError:
                pass
                
    def _store_demographics_entry_widgets(self):
        """Store references to entry widgets that should be disabled when name is empty."""
        self.demographics_entry_widgets = []
        
    def _is_demographics_form_empty(self) -> bool:
        """Check if the demographics form is empty.
        
        Returns:
            True if all fields are empty, False otherwise
        """
        if self.vars.first_name.get().strip():
            return False
        if self.vars.last_name.get().strip():
            return False
        if self.vars.address.get().strip():
            return False
        if self.vars.cell_phone.get().strip():
            return False
        if self.vars.home_phone.get().strip():
            return False
        if self.vars.email.get().strip():
            return False
        if self.vars.alternate_email.get().strip():
            return False
        if self.vars.emergency_contact_name.get().strip():
            return False
        if self.vars.emergency_contact_phone.get().strip():
            return False
        if self.vars.ham_callsign.get().strip():
            return False
        if self.vars.mission_eligible.get():
            return False
        return True
    
    def _on_member_status_changed(self):
        """Handle member status radiobutton change.
        
        Auto-saves the status if a member is currently being edited.
        """
        member_id = self.vars.selected_member_id.get()
        
        if member_id > 0 and self.db.database_exists():
            # Auto-save the status change
            member = self.db.get_member(member_id)
            if member:
                member['member_status'] = self.vars.member_status.get()
                self.db.save_member(member)
                
                # Refresh the members list and attendance tree to show updated status
                self.ui_state.members_list = self.db.get_all_members()
                self._refresh_members_tree()
                self._refresh_attendance_tree()
                
    def _on_new_member(self):
        """Handle New Member button click."""
        # Skip confirmation if form is already empty
        if not self._is_demographics_form_empty():
            if not messagebox.askyesno("New Member", 
                                       "Clear form for new member entry?\n"
                                       "Any unsaved changes will be lost."):
                return
            
        self.vars.clear_demographics()
        self._update_cert_tree_from_vars()
        self.ui_state.is_editing_member = False
        self._update_demographics_entry_states()
        
    def _on_save_member(self):
        """Handle Save Member button click."""
        # Collect form data
        form_data = {
            "first_name": self.vars.first_name.get(),
            "last_name": self.vars.last_name.get(),
            "email": self.vars.email.get(),
            "alternate_email": self.vars.alternate_email.get(),
            "cell_phone": self.vars.cell_phone.get(),
            "home_phone": self.vars.home_phone.get(),
            "emergency_contact_phone": self.vars.emergency_contact_phone.get(),
        }
        
        # Validate
        is_valid, errors = ui_support.validate_demographics_form(form_data)
        if not is_valid:
            messagebox.showerror("Validation Error", "\n".join(errors))
            return
            
        if not self.db.database_exists():
            messagebox.showwarning("Save Member", 
                                   "Please initialize the database first (Setup tab).")
            return
            
        # Prepare full member data (title case for names)
        member_data = {
            'id': self.vars.selected_member_id.get(),
            'first_name': self.vars.first_name.get().strip().title(),
            'last_name': self.vars.last_name.get().strip().title(),
            'address': self.vars.address.get().strip(),
            'cell_phone': self.vars.cell_phone.get().strip(),
            'home_phone': self.vars.home_phone.get().strip(),
            'email': self.vars.email.get().strip(),
            'alternate_email': self.vars.alternate_email.get().strip(),
            'emergency_contact_name': self.vars.emergency_contact_name.get().strip().title(),
            'emergency_contact_phone': self.vars.emergency_contact_phone.get().strip(),
            'ham_callsign': self.vars.ham_callsign.get().strip(),
            'mission_eligible': self.vars.mission_eligible.get(),
            'member_status': self.vars.member_status.get(),
            'certifications': self._get_certifications_from_tree(),
        }
        
        # Update the UI fields to show the title case
        self.vars.first_name.set(member_data['first_name'])
        self.vars.last_name.set(member_data['last_name'])
        self.vars.emergency_contact_name.set(member_data['emergency_contact_name'])
        
        # Save to database
        success, member_id, message = self.db.save_member(member_data)
        
        if success:
            self.vars.selected_member_id.set(member_id)
            
            # Refresh member lists
            self.ui_state.members_list = self.db.get_all_members()
            self._refresh_members_tree()
            self.refresh_last_name_combobox()
            
            # Initialize attendance records for this member in existing sessions
            self._ensure_member_attendance_records(member_id)
            
            # Refresh attendance tree to show the new member
            self._refresh_attendance_tree()
            
            # Clear form for new member entry (Save implies New)
            self.vars.clear_demographics()
            self._update_cert_tree_from_vars()
            self.ui_state.is_editing_member = False
            self._update_demographics_entry_states()
        else:
            messagebox.showerror("Save Member", message)
            
    def _ensure_member_attendance_records(self, member_id: int):
        """Ensure attendance records exist for a member in all sessions.
        
        Only creates records for sessions that don't already have one.
        Does NOT overwrite existing attendance data.
        
        Args:
            member_id: ID of the member
        """
        if not self.db.database_exists():
            return
            
        try:
            conn = self.db._get_connection()
            cursor = conn.cursor()
            
            # Get all session IDs
            cursor.execute("SELECT id FROM training_sessions")
            session_ids = [row[0] for row in cursor.fetchall()]
            
            # Insert attendance records only where they don't exist (INSERT OR IGNORE)
            for session_id in session_ids:
                cursor.execute('''
                    INSERT OR IGNORE INTO attendance (session_id, member_id, attended)
                    VALUES (?, ?, 0)
                ''', (session_id, member_id))
                
            conn.commit()
            
        except Exception as e:
            print(f"Error ensuring attendance records: {e}")
        
    def _on_delete_member(self):
        """Handle Delete Member button click."""
        member_id = self.vars.selected_member_id.get()
        
        if member_id < 0:
            # Try to get from tree selection
            selection = self.members_tree.selection()
            if not selection:
                messagebox.showwarning("Delete Member", "Please select a member to delete.")
                return
            # Extract member_id from item id
            try:
                member_id = int(selection[0].replace("member_", ""))
            except ValueError:
                messagebox.showwarning("Delete Member", "Please select a member to delete.")
                return
        
        if not self.db.database_exists():
            messagebox.showwarning("Delete Member", 
                                   "Please initialize the database first (Setup tab).")
            return
            
        # Get member name for confirmation message
        member = self.db.get_member(member_id)
        if not member:
            messagebox.showwarning("Delete Member", "Member not found.")
            return
            
        member_name = f"{member['first_name']} {member['last_name']}"
            
        if not messagebox.askyesno("Confirm Delete", 
                                   f"Are you sure you want to delete {member_name}?\n"
                                   "This action cannot be undone."):
            return
            
        success, message = self.db.delete_member(member_id)
        
        if success:
            self.vars.clear_demographics()
            self._update_cert_tree_from_vars()
            
            # Refresh member lists
            self.ui_state.members_list = self.db.get_all_members()
            self._refresh_members_tree()
            self.refresh_last_name_combobox()
            
            # Refresh attendance tree to remove the deleted member
            self._refresh_attendance_tree()
            self._update_demographics_entry_states()
        else:
            messagebox.showerror("Delete Member", message)
            
    def _on_clear_demographics(self):
        """Handle Clear Form button click."""
        # Skip confirmation if form is already empty
        if not self._is_demographics_form_empty():
            if not messagebox.askyesno("Clear Form", 
                                       "Clear all form fields?\n"
                                       "Any unsaved changes will be lost."):
                return
            
        self.vars.clear_demographics()
        self._update_cert_tree_from_vars()
        self.ui_state.is_editing_member = False
        self._update_demographics_entry_states()
        
    def _on_member_selected(self, event):
        """Handle member selection in treeview.
        
        Args:
            event: The tkinter event
        """
        selection = self.members_tree.selection()
        if selection:
            try:
                member_id = int(selection[0].replace("member_", ""))
                member = self.db.get_member(member_id)
                if member:
                    self._load_member_data(member)
            except (ValueError, AttributeError):
                pass
    
    def _on_last_name_selected(self, event):
        """Handle last name selection from combobox dropdown.
        
        Args:
            event: The tkinter event
        """
        selected_name = self.vars.last_name.get()
        if selected_name and self.db.database_exists():
            # Find members with this last name
            members = self.db.get_all_members()
            matching = [m for m in members if m.get("last_name", "") == selected_name]
            
            if len(matching) == 1:
                self._load_member_data(matching[0])
            elif len(matching) > 1:
                # Multiple members with same last name - show selection
                # For now, load the first one
                self._load_member_data(matching[0])
                    
    def _load_member_data(self, member: dict):
        """Load member data into the form fields.
        
        Args:
            member: Dictionary containing member data
        """
        # Set flag to prevent duplicate name warning while loading
        self._loading_member_data = True
        
        self.vars.first_name.set(member.get("first_name", ""))
        self.vars.last_name.set(member.get("last_name", ""))
        self.vars.address.set(member.get("address", ""))
        self.vars.cell_phone.set(member.get("cell_phone", ""))
        self.vars.home_phone.set(member.get("home_phone", ""))
        self.vars.email.set(member.get("email", ""))
        self.vars.alternate_email.set(member.get("alternate_email", ""))
        self.vars.emergency_contact_name.set(member.get("emergency_contact_name", ""))
        self.vars.emergency_contact_phone.set(member.get("emergency_contact_phone", ""))
        self.vars.ham_callsign.set(member.get("ham_callsign", ""))
        self.vars.mission_eligible.set(member.get("mission_eligible", False))
        self.vars.member_status.set(member.get("member_status", "Member"))
        self.vars.selected_member_id.set(member.get("id", -1))
        
        # Load certification dates
        certs = member.get("certifications", {})
        self.vars.set_certification_dates_from_dict(certs)
        self._update_cert_tree_from_vars()
        
        self.ui_state.is_editing_member = True
        self._loading_member_data = False
        
        # Update entry states (should be enabled now that name is filled)
        self._update_demographics_entry_states()
        
    def refresh_last_name_combobox(self):
        """Refresh the last name combobox with current members, sorted alphabetically."""
        if self.db.database_exists():
            last_names = self.db.get_unique_last_names()
        else:
            last_names = sorted(set(
                member.get("last_name", "") 
                for member in self.ui_state.members_list 
                if member.get("last_name", "")
            ), key=str.lower)
        
        self.last_name_combo['values'] = last_names
        
    def _refresh_members_tree(self):
        """Refresh the members treeview from current data."""
        # Clear current items
        for item in self.members_tree.get_children():
            self.members_tree.delete(item)
            
        # Add members with alternating colors
        for i, member in enumerate(self.ui_state.members_list):
            name = ui_support.format_member_display_name(
                member.get("first_name", ""),
                member.get("last_name", "")
            )
            callsign = member.get("ham_callsign", "")
            eligible = "Yes" if member.get("mission_eligible", False) else "No"
            cell_phone = member.get("cell_phone", "")
            
            tag = 'oddrow' if i % 2 else 'evenrow'
            self.members_tree.insert("", tk.END,
                                    values=(name, callsign, eligible, cell_phone),
                                    iid=f"member_{member.get('id', 0)}",
                                    tags=(tag,))
            
    def _update_cert_tree_from_vars(self):
        """Update certification treeview from member's certification data."""
        # Get current certification dates from vars
        cert_dates = self.vars.get_certification_dates_dict()
        
        # Update each row in the treeview
        for item in self.cert_tree.get_children():
            values = self.cert_tree.item(item, "values")
            if values:
                cert_name = values[0]
                date_value = cert_dates.get(cert_name, "")
                self.cert_tree.item(item, values=(cert_name, date_value))
                
    def _get_certifications_from_tree(self) -> dict:
        """Get certification dates from the treeview.
        
        Returns:
            Dictionary with certification names as keys and dates as values
        """
        cert_dates = {}
        for item in self.cert_tree.get_children():
            values = self.cert_tree.item(item, "values")
            if values and len(values) >= 2:
                cert_name = values[0]
                date_value = values[1] if values[1] else ""
                cert_dates[cert_name] = date_value
        return cert_dates
            
    # ============================================================
    # Event Handlers - Training Sessions Tab
    # ============================================================
    
    def _get_sessions_for_calendar(self) -> list:
        """Get sessions list for calendar highlighting.
        
        Returns:
            List of session dictionaries
        """
        return self.ui_state.sessions_list
        
    def _on_session_date_selected(self, event=None):
        """Handle date selection from datepicker.
        
        Args:
            event: The tkinter event
        """
        if ui_support.TKCALENDAR_AVAILABLE:
            # Get the selected date and format it
            selected_date = self.date_entry.get_date()
            date_str = selected_date.strftime(ui_support.DATE_FORMAT)
            self.vars.session_date.set(date_str)
        else:
            date_str = self.vars.session_date.get()
        
        # Update attendance section state based on whether date is selected
        self._update_attendance_section_state()
        
        # Only auto-set Weekend Training/Weeknight Training if a date was actually selected (not empty)
        if date_str and date_str.strip():
            suggested_type = ui_support.suggest_session_type(date_str)
            current_type = self.vars.session_type.get()
            
            # Only auto-set if currently Weekend Training or Weeknight Training (not Mission/Other)
            if current_type in ("Weekend Training", "Weeknight Training"):
                self.vars.session_type.set(suggested_type)
                self._on_session_type_changed()
                
            # Check if there's an existing session on this date and load it
            self._check_and_load_existing_session(date_str)
            
    def _update_attendance_section_state(self):
        """Enable/disable attendance section and radio buttons based on whether a date is selected."""
        date_str = self.vars.session_date.get().strip()
        has_date = bool(date_str)
        
        if has_date:
            # Enable attendance section
            self.attendance_tree.configure(selectmode='browse')
            self.attendance_tree.bind("<Double-1>", self._on_attendance_double_click)
            # Update instruction text
            self.attendance_instructions.configure(text="Double-click 'Attended' column to toggle Yes/No")
            # Enable radio buttons
            for rb in self.session_type_radios:
                rb.configure(state='normal')
        else:
            # Disable attendance section
            self.attendance_tree.configure(selectmode='none')
            self.attendance_tree.unbind("<Double-1>")
            # Update instruction text to indicate date needed
            self.attendance_instructions.configure(text="Select a date to enable attendance tracking")
            # Disable radio buttons
            for rb in self.session_type_radios:
                rb.configure(state='disabled')
        
    def _on_session_date_focus_out(self, event):
        """Handle focus out on session date entry (fallback for non-datepicker).
        
        Args:
            event: The tkinter event
        """
        date_str = self.vars.session_date.get()
        if date_str:
            is_valid, result = ui_support.validate_date(date_str)
            if is_valid:
                self.vars.session_date.set(result)
                # Auto-set Weekend Training/Weeknight Training
                suggested_type = ui_support.suggest_session_type(result)
                current_type = self.vars.session_type.get()
                if current_type in ("Weekend Training", "Weeknight Training"):
                    self.vars.session_type.set(suggested_type)
                    self._on_session_type_changed()
                # Check for existing session
                self._check_and_load_existing_session(result)
            else:
                messagebox.showwarning("Date Validation Error", result)
                event.widget.focus_set()
                
    def _on_session_type_changed(self):
        """Handle session type radio button change."""
        session_type = self.vars.session_type.get()
        
        # Show/hide description field based on type
        if session_type in ("Mission", "Team Meeting", "Other"):
            self.description_frame.pack(fill=tk.X, pady=(5, 0))
        else:
            self.description_frame.pack_forget()
            self.vars.session_description.set("")
        
        # Skip the dialog if we're loading session data
        if self._loading_session_data:
            return
        
        # If editing an existing session, ask whether to update or create new
        session_id = self.vars.selected_session_id.get()
        if session_id > 0 and self.db.database_exists():
            # Get current session data
            location = self.vars.session_location.get().strip()
            date_str = self.vars.session_date.get().strip()
            description = self.vars.session_description.get().strip()
            
            if location and date_str:
                response = messagebox.askyesnocancel(
                    "Session Type Changed",
                    f"You changed the session type to '{session_type}'.\n\n"
                    "â€¢ Yes - Update the existing session\n"
                    "â€¢ No - Create a new session with this type\n"
                    "â€¢ Cancel - Revert to previous type"
                )
                
                if response is True:  # Yes - Update existing
                    session_data = {
                        'id': session_id,
                        'location': location,
                        'date': date_str,
                        'type': session_type,
                        'description': description,
                    }
                    
                    success, _, _ = self.db.save_session(session_data)
                    
                    if success:
                        # Track the year as modified for Excel export
                        try:
                            session_date = datetime.strptime(date_str, "%m/%d/%Y")
                            self._modified_years.add(session_date.year)
                        except ValueError:
                            pass
                        
                        # Refresh session lists
                        self.ui_state.sessions_list = self.db.get_all_sessions()
                        
                elif response is False:  # No - Create new session
                    # Clear the session ID so it will be saved as a new session
                    self.vars.selected_session_id.set(-1)
                    self.ui_state.is_editing_session = False
                    # Description is cleared for new session unless it's a type that uses it
                    if session_type not in ("Mission", "Team Meeting", "Other"):
                        self.vars.session_description.set("")
                    
                else:  # Cancel - Revert
                    # Get the original session type from database
                    original_session = self.db.get_session(session_id)
                    if original_session:
                        original_type = original_session.get('type', 'Weekend Training')
                        self.vars.session_type.set(original_type)
                        # Update description field visibility for original type
                        if original_type in ("Mission", "Team Meeting", "Other"):
                            self.description_frame.pack(fill=tk.X, pady=(5, 0))
                        else:
                            self.description_frame.pack_forget()
            
    def _check_and_load_existing_session(self, date_str: str):
        """Check if there are existing sessions on the date and inform the user.
        
        Multiple sessions can exist on the same day at different locations,
        or with different types (e.g., training + mission).
        
        Args:
            date_str: Date string in MM/DD/YYYY format
        """
        if not self.db.database_exists():
            return
            
        matching_sessions = self.db.get_sessions_by_date(date_str)
        
        if not matching_sessions:
            return
            
        if len(matching_sessions) == 1:
            # Only one session exists on this date
            session = matching_sessions[0]
            session_type = session.get('type', 'Unknown')
            location = session.get('location', '')
            response = messagebox.askyesnocancel(
                "Existing Session Found",
                f"A {session_type} session already exists for {date_str} at {location}.\n\n"
                "\u2022 Yes - Load the existing session\n"
                "\u2022 No - Create a new session (different location or type)\n"
                "\u2022 Cancel - Clear the date"
            )
            if response is True:  # Yes - load it
                self._load_session_data(session)
            elif response is False:  # No - continue with new session
                pass  # User wants to create a new session
            else:  # Cancel
                self.vars.session_date.set("")
                if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
                    self.date_entry.delete(0, tk.END)
        else:
            # Multiple sessions exist on this date
            self._show_session_selection_popup(date_str, matching_sessions)
            
    def _show_session_selection_popup(self, date_str: str, sessions: list):
        """Show a popup to select which session to load when multiple exist.
        
        Args:
            date_str: Date string
            sessions: List of session dictionaries
        """
        popup = tk.Toplevel(self.root)
        popup.title(f"Sessions on {date_str}")
        popup.geometry("500x350")
        popup.transient(self.root)
        popup.grab_set()
        
        # Center on parent
        popup.geometry(f"+{self.root.winfo_x() + 150}+{self.root.winfo_y() + 150}")
        
        # Instructions
        ttk.Label(popup, 
                  text=f"Multiple sessions exist for {date_str}.\n"
                       "Select one to load, or click 'New Session' to create another:",
                  justify=tk.LEFT).pack(pady=(10, 5), padx=10, anchor=tk.W)
        
        # Listbox with scrollbar
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, height=8)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox
        session_map = {}
        for session in sessions:
            session_type = session.get('type', 'Unknown')
            location = session.get('location', '')
            description = session.get('description', '')
            display_text = f"{session_type} at {location}"
            if description:
                display_text += f" - {description[:30]}..."
            listbox.insert(tk.END, display_text)
            session_map[display_text] = session
            
        # Button frame
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def on_load():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Select Session", "Please select a session to load.")
                return
                
            display_text = listbox.get(selection[0])
            session = session_map.get(display_text)
            
            if session:
                popup.destroy()
                self._load_session_data(session)
                
        def on_new():
            popup.destroy()
            # User wants to create a new session on this date
            
        def on_cancel():
            popup.destroy()
            self.vars.session_date.set("")
            if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
                self.date_entry.delete(0, tk.END)
            
        ttk.Button(btn_frame, text="Load Selected", command=on_load).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="New Session", command=on_new).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.LEFT, padx=5)
                    
    def _load_session_data(self, session: dict):
        """Load session data into the form fields.
        
        Args:
            session: Dictionary containing session data
        """
        # Set flag to prevent session type change dialog while loading
        self._loading_session_data = True
        
        self.vars.session_location.set(session.get('location', ''))
        self.vars.session_date.set(session.get('date', ''))
        self.vars.session_type.set(session.get('type', 'Weekend Training'))
        self.vars.session_description.set(session.get('description', ''))
        self.vars.selected_session_id.set(session.get('id', -1))
        
        # Update description field visibility
        self._on_session_type_changed()
        
        # Clear the loading flag
        self._loading_session_data = False
        
        # Load attendance data from database
        session_id = session.get('id', -1)
        if session_id > 0 and self.db.database_exists():
            self._refresh_attendance_tree()
        
        self.ui_state.is_editing_session = True
        
        # Mark this as the saved state for change tracking
        self.ui_state.mark_session_saved(self._get_current_session_data())
        
    def _update_attendance_from_state(self):
        """Update attendance treeview from current session attendance state."""
        for item in self.attendance_tree.get_children():
            # Extract member_id from item id (format: "member_{id}")
            try:
                member_id = int(item.replace("member_", ""))
                attendance = self.ui_state.current_session_attendance.get(member_id, "No")
                values = list(self.attendance_tree.item(item, "values"))
                values[1] = "Yes" if attendance else "No"
                self.attendance_tree.item(item, values=values)
            except (ValueError, IndexError):
                pass
    
    def refresh_location_combobox(self):
        """Refresh the location combobox with existing locations."""
        if self.db.database_exists():
            # Get locations from both training_locations table and used locations
            default_locations = self.db.get_training_locations()
            session_locations = ui_support.get_unique_locations(self.ui_state.sessions_list)
            
            # Combine and sort
            all_locations = sorted(set(default_locations + session_locations), key=str.lower)
            self.location_combo['values'] = all_locations
        else:
            locations = ui_support.get_unique_locations(self.ui_state.sessions_list)
            self.location_combo['values'] = locations
    
    def _on_new_session(self):
        """Handle New Session button click."""
        if not messagebox.askyesno("New Session", 
                                   "Clear form for new session entry?"):
            return
            
        self.vars.clear_training_session()
        self.ui_state.is_editing_session = False
        self.ui_state.current_session_attendance = {}
        self._reset_attendance_list()
        self._on_session_type_changed()  # Hide description field
        
        # Clear the datepicker if using tkcalendar
        if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
            self.date_entry.delete(0, tk.END)
            
        # Mark as clean (no unsaved changes)
        self.ui_state.mark_session_clean()
        
        # Update attendance section state (disabled since no date)
        self._update_attendance_section_state()
        
    def _on_save_session(self):
        """Handle Save Session button click."""
        # Collect form data
        form_data = {
            "location": self.vars.session_location.get(),
            "date": self.vars.session_date.get(),
            "type": self.vars.session_type.get(),
        }
        
        # Validate
        is_valid, errors = ui_support.validate_training_session_form(form_data)
        if not is_valid:
            messagebox.showerror("Validation Error", "\n".join(errors))
            return
            
        if not self.db.database_exists():
            messagebox.showwarning("Save Session", 
                                   "Please initialize the database first (Setup tab).")
            return
        
        # Check for description if Mission/Other
        session_type = self.vars.session_type.get()
        description = self.vars.session_description.get().strip()
        
        if session_type in ("Mission", "Other") and not description:
            if not messagebox.askyesno("Missing Description",
                                      f"No description provided for {session_type} session.\n"
                                      "Continue saving without a description?"):
                return
        
        # Auto-add new location to training locations list (with title case)
        location = self.vars.session_location.get().strip().title()
        
        # Update the UI field to show the title case
        self.vars.session_location.set(location)
        
        existing_locations = self.db.get_training_locations()
        if location and location not in existing_locations:
            self.db.add_training_location(location)
            self._refresh_locations_listbox()
                
        # Prepare session data
        session_data = {
            'id': self.vars.selected_session_id.get(),
            'location': location,
            'date': self.vars.session_date.get().strip(),
            'type': session_type,
            'description': description,
        }
        
        # Save to database
        success, session_id, message = self.db.save_session(session_data)
        
        if success:
            self.vars.selected_session_id.set(session_id)
            
            # Save attendance records
            self._save_current_attendance(session_id)
            
            # Track the year as modified for Excel export
            date_str = self.vars.session_date.get()
            if date_str:
                try:
                    session_date = datetime.strptime(date_str, "%m/%d/%Y")
                    self._modified_years.add(session_date.year)
                except ValueError:
                    pass
            
            # Refresh session lists
            self.ui_state.sessions_list = self.db.get_all_sessions()
            self.refresh_location_combobox()
            
            # Mark session as saved (no unsaved changes)
            self.ui_state.mark_session_clean()
            
            # Clear form for new session entry (Save implies New)
            self.vars.clear_training_session()
            self._on_session_type_changed()  # Hide description field
            self._refresh_attendance_tree()
            
            # Clear the datepicker if using tkcalendar
            if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
                self.date_entry.delete(0, tk.END)
                
            # Update attendance section state (disabled since no date)
            self._update_attendance_section_state()
        else:
            messagebox.showerror("Save Session", message)
            
    def _save_current_attendance(self, session_id: int):
        """Save the current attendance from the treeview to database.
        
        Args:
            session_id: ID of the session
        """
        for item in self.attendance_tree.get_children():
            try:
                member_id = int(item.replace("member_", ""))
                values = self.attendance_tree.item(item, "values")
                attended = values[1] == "Yes" if len(values) > 1 else False
                self.db.update_attendance(session_id, member_id, attended)
            except (ValueError, IndexError):
                continue
        
    def _on_delete_session(self):
        """Handle Delete Session button click."""
        session_id = self.vars.selected_session_id.get()
        
        if session_id < 0:
            messagebox.showwarning("Delete Session", "No session is currently selected.")
            return
            
        if not self.db.database_exists():
            messagebox.showwarning("Delete Session", 
                                   "Please initialize the database first (Setup tab).")
            return
        
        # Get session info for confirmation message
        date_str = self.vars.session_date.get()
        location = self.vars.session_location.get()
        session_type = self.vars.session_type.get()
            
        if not messagebox.askyesno("Confirm Delete", 
                                   f"Are you sure you want to delete this session?\n\n"
                                   f"Date: {date_str}\n"
                                   f"Location: {location}\n"
                                   f"Type: {session_type}\n\n"
                                   "This action cannot be undone."):
            return
            
        success, message = self.db.delete_session(session_id)
        
        if success:
            # Track the year as modified for Excel export
            if date_str:
                try:
                    session_date = datetime.strptime(date_str, "%m/%d/%Y")
                    self._modified_years.add(session_date.year)
                except ValueError:
                    pass
            
            self.vars.clear_training_session()
            self.ui_state.current_session_attendance = {}
            self.ui_state.is_editing_session = False
            self._on_session_type_changed()
            
            # Clear the datepicker if using tkcalendar
            if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
                self.date_entry.delete(0, tk.END)
            
            # Mark session as clean (no unsaved changes)
            self.ui_state.mark_session_clean()
            
            # Update attendance section state (disabled since no date)
            self._update_attendance_section_state()
            
            # Refresh session lists
            self.ui_state.sessions_list = self.db.get_all_sessions()
            self.refresh_location_combobox()
            self._refresh_attendance_tree()
        else:
            messagebox.showerror("Delete Session", message)
            
    def _on_clear_session(self):
        """Handle Clear Form button click."""
        if not messagebox.askyesno("Clear Form", 
                                   "Clear all form fields?"):
            return
            
        self.vars.clear_training_session()
        self.ui_state.current_session_attendance = {}
        self._reset_attendance_list()
        self._on_session_type_changed()  # Hide description field
        
        # Clear the datepicker if using tkcalendar
        if ui_support.TKCALENDAR_AVAILABLE and self.date_entry:
            self.date_entry.delete(0, tk.END)
            
        # Mark as clean (no unsaved changes)
        self.ui_state.mark_session_clean()
        
        # Update attendance section state (disabled since no date)
        self._update_attendance_section_state()
        
    def _on_modify_session(self):
        """Handle Modify Date and/or Location button click.
        
        Updates the date and/or location of the currently selected session.
        """
        session_id = self.vars.selected_session_id.get()
        
        if session_id < 0:
            messagebox.showwarning("No Session Selected", 
                                   "Please select an existing session to modify.\n\n"
                                   "Use the date picker to select a session first.")
            return
            
        if not self.db.database_exists():
            messagebox.showwarning("Database Required", 
                                   "Please initialize the database first (Setup tab).")
            return
        
        # Get current form values
        new_location = self.vars.session_location.get().strip().title()
        new_date = self.vars.session_date.get().strip()
        session_type = self.vars.session_type.get()
        description = self.vars.session_description.get().strip()
        
        # Validate required fields
        if not new_location or not new_date:
            messagebox.showwarning("Missing Information", 
                                   "Please enter both Location and Date.")
            return
        
        # Validate date format
        is_valid, result = ui_support.validate_date(new_date)
        if not is_valid:
            messagebox.showwarning("Invalid Date", result)
            return
        new_date = result
        
        # Get original session info for confirmation message
        original_session = self.db.get_session(session_id)
        if not original_session:
            messagebox.showerror("Error", "Could not find the original session.")
            return
        
        original_date = original_session.get('date', '')
        original_location = original_session.get('location', '')
        
        # Check if anything actually changed
        if new_date == original_date and new_location == original_location:
            messagebox.showinfo("No Changes", 
                               "The date and location are the same as the original.\n"
                               "No changes were made.")
            return
        
        # Confirm the modification
        changes = []
        if new_date != original_date:
            changes.append(f"Date: {original_date} Ã¢â€ â€™ {new_date}")
        if new_location != original_location:
            changes.append(f"Location: {original_location} Ã¢â€ â€™ {new_location}")
        
        change_text = "\n".join(changes)
        
        if not messagebox.askyesno("Confirm Modification",
                                   f"Modify this session?\n\n{change_text}\n\n"
                                   "All attendance records will be preserved."):
            return
        
        # Update the UI field to show title case
        self.vars.session_location.set(new_location)
        self.vars.session_date.set(new_date)
        
        # Auto-add new location to training locations list if needed
        existing_locations = self.db.get_training_locations()
        if new_location and new_location not in existing_locations:
            self.db.add_training_location(new_location)
            self._refresh_locations_listbox()
        
        # Prepare session data for update
        session_data = {
            'id': session_id,
            'location': new_location,
            'date': new_date,
            'type': session_type,
            'description': description,
        }
        
        # Save to database
        success, updated_id, message = self.db.save_session(session_data)
        
        if success:
            # Track the year(s) as modified for Excel export
            for date_str in [original_date, new_date]:
                if date_str:
                    try:
                        session_date = datetime.strptime(date_str, "%m/%d/%Y")
                        self._modified_years.add(session_date.year)
                    except ValueError:
                        pass
            
            # Refresh session lists
            self.ui_state.sessions_list = self.db.get_all_sessions()
            self.refresh_location_combobox()
            
            # Verify the update by re-reading from database
            updated_session = self.db.get_session(session_id)
            if updated_session:
                actual_location = updated_session.get('location', '')
                actual_date = updated_session.get('date', '')
                if actual_location != new_location or actual_date != new_date:
                    messagebox.showwarning("Verification Warning",
                        f"Session may not have saved correctly.\n\n"
                        f"Expected location: {new_location}\n"
                        f"Actual location: {actual_location}\n\n"
                        f"Expected date: {new_date}\n"
                        f"Actual date: {actual_date}")
            
            # Mark session as saved
            self.ui_state.mark_session_saved(self._get_current_session_data())
            
            messagebox.showinfo("Session Modified", 
                               f"Session successfully updated.\n\n{change_text}")
        else:
            messagebox.showerror("Modification Failed", message)
        
    def _on_attendance_double_click(self, event):
        """Handle double-click on attendance treeview to toggle yes/no.
        
        Auto-saves the session if not already saved.
        
        Args:
            event: The tkinter event
        """
        # Get the item and column that was clicked
        region = self.attendance_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
            
        column = self.attendance_tree.identify_column(event.x)
        item = self.attendance_tree.identify_row(event.y)
        
        # Attended column is now #3 (after name and status)
        if not item or column != "#3":  # Only allow toggling the attended column
            return
        
        if not self.db.database_exists():
            messagebox.showwarning("Database Required", 
                                   "Please initialize the database first (Setup tab).")
            return
            
        # Check if we need to save the session first
        session_id = self.vars.selected_session_id.get()
        
        if session_id < 0:
            # Session not saved yet - check if we have required fields
            location = self.vars.session_location.get().strip()
            date_str = self.vars.session_date.get().strip()
            
            if not location or not date_str:
                messagebox.showwarning("Session Required", 
                                       "Please enter a Location and Date before marking attendance.")
                return
            
            # Auto-save the session
            session_id = self._auto_save_session()
            if session_id < 0:
                # Save failed
                return
        
        # Toggle the value (attended is now at index 2: name, status, attended)
        values = list(self.attendance_tree.item(item, "values"))
        current_attendance = values[2]
        new_attendance = "Yes" if current_attendance == "No" else "No"
        values[2] = new_attendance
        self.attendance_tree.item(item, values=values)
        
        # Update database
        try:
            member_id = int(item.replace("member_", ""))
            attended = new_attendance == "Yes"
            self.db.update_attendance(session_id, member_id, attended)
            
            # Track the year as modified for Excel export
            date_str = self.vars.session_date.get()
            if date_str:
                try:
                    session_date = datetime.strptime(date_str, "%m/%d/%Y")
                    self._modified_years.add(session_date.year)
                except ValueError:
                    pass
            
            # COMMENTED OUT: Update weekend count display
            # weekend_count = self.db.get_weekend_attendance_count(member_id)
            # values[3] = str(weekend_count)
            # self.attendance_tree.item(item, values=values)
        except ValueError:
            pass
            
    def _auto_save_session(self) -> int:
        """Auto-save the current session to database.
        
        Returns:
            Session ID if successful, -1 if failed
        """
        # Get and validate form data
        location = self.vars.session_location.get().strip().title()
        date_str = self.vars.session_date.get().strip()
        session_type = self.vars.session_type.get()
        description = self.vars.session_description.get().strip()
        
        # Validate date
        is_valid, result = ui_support.validate_date(date_str)
        if not is_valid:
            messagebox.showerror("Invalid Date", result)
            return -1
        date_str = result
        
        # Update UI with normalized values
        self.vars.session_location.set(location)
        self.vars.session_date.set(date_str)
        
        # Auto-add new location to training locations list
        existing_locations = self.db.get_training_locations()
        if location and location not in existing_locations:
            self.db.add_training_location(location)
            self._refresh_locations_listbox()
        
        # Prepare session data
        session_data = {
            'id': self.vars.selected_session_id.get(),
            'location': location,
            'date': date_str,
            'type': session_type,
            'description': description,
        }
        
        # Save to database
        success, session_id, message = self.db.save_session(session_data)
        
        if success:
            self.vars.selected_session_id.set(session_id)
            
            # Track the year as modified for Excel export
            try:
                session_date = datetime.strptime(date_str, "%m/%d/%Y")
                self._modified_years.add(session_date.year)
            except ValueError:
                pass
            
            # Refresh session lists
            self.ui_state.sessions_list = self.db.get_all_sessions()
            self.refresh_location_combobox()
            
            # Mark session as saved
            self.ui_state.mark_session_saved(self._get_current_session_data())
            
            return session_id
        else:
            messagebox.showerror("Save Error", message)
            return -1
        
    def _reset_attendance_list(self):
        """Reset the attendance list to default (all No)."""
        for item in self.attendance_tree.get_children():
            values = list(self.attendance_tree.item(item, "values"))
            # Attended is now column index 2 (name=0, status=1, attended=2)
            values[2] = "No"
            self.attendance_tree.item(item, values=values)
            
    def _refresh_attendance_tree(self):
        """Refresh the attendance treeview with current members.
        
        Members are sorted by status (Member, Candidate, Guest) first,
        then by name within each status group.
        """
        # Clear current items
        for item in self.attendance_tree.get_children():
            self.attendance_tree.delete(item)
            
        if not self.db.database_exists():
            return
        
        # Get sort preference
        sort_by_first = getattr(self, 'attendance_sort_var', None) and self.attendance_sort_var.get() == "first"
            
        session_id = self.vars.selected_session_id.get()
        
        if session_id > 0:
            # Get attendance summary for this session
            attendance_data = self.db.get_member_attendance_summary(session_id)
            
            # Sort by status first, then by name using the helper function
            attendance_data = ui_support.sort_members_by_status_then_name(attendance_data, sort_by_first)
            
            for i, data in enumerate(attendance_data):
                # Format name based on sort preference
                if sort_by_first:
                    name = f"{data['first_name']} {data['last_name']}"
                else:
                    name = f"{data['last_name']}, {data['first_name']}"
                
                status = data.get('member_status', 'Member')
                attended = "Yes" if data['attended'] else "No"
                # COMMENTED OUT: weekend_count = str(data['weekend_count'])
                
                tag = 'oddrow' if i % 2 else 'evenrow'
                self.attendance_tree.insert("", tk.END,
                                           values=(name, status, attended),
                                           # COMMENTED OUT: values=(name, status, attended, weekend_count),
                                           iid=f"member_{data['id']}",
                                           tags=(tag,))
        else:
            # No session selected, show all members with default No
            members = self.db.get_all_members()
            
            # Sort by status first, then by name using the helper function
            members = ui_support.sort_members_by_status_then_name(members, sort_by_first)
            
            for i, member in enumerate(members):
                # Format name based on sort preference
                if sort_by_first:
                    name = f"{member.get('first_name', '')} {member.get('last_name', '')}"
                else:
                    name = f"{member.get('last_name', '')}, {member.get('first_name', '')}"
                
                status = member.get('member_status', 'Member')
                # COMMENTED OUT: weekend_count = self.db.get_weekend_attendance_count(member['id'])
                
                tag = 'oddrow' if i % 2 else 'evenrow'
                self.attendance_tree.insert("", tk.END,
                                           values=(name, status, "No"),
                                           # COMMENTED OUT: values=(name, status, "No", str(weekend_count)),
                                           iid=f"member_{member['id']}",
                                           tags=(tag,))
            
    def refresh_attendance_list(self):
        """Refresh the attendance list from the members list."""
        self._refresh_attendance_tree()


def main():
    """Main entry point for the application."""
    root = tk.Tk()
    
    # Set a nicer theme if available
    style = ttk.Style()
    available_themes = style.theme_names()
    if 'clam' in available_themes:
        style.theme_use('clam')
    elif 'vista' in available_themes:
        style.theme_use('vista')

    # Read saved window geometry from config file for splash screen positioning
    saved_geometry = None
    try:
        cfg = config.load_config()
        saved_geometry = cfg.get("window_geometry", "")
    except Exception:
        pass  # If config can't be loaded, splash will center on screen
    
    # Show splash screen centered over saved main window position
    splash = SplashScreen(root, version="1.0.4-alpha",
                          app_title="Attendance Tracker", 
                          github_url="github.com/agelders2021/attendance-tracker",
                          main_window_geometry=saved_geometry)
    app = TrainingTrackerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
